import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict
import tempfile

class ExcelReportGenerator:
    def __init__(self):
        self.header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_report(self, report_type, data, filters):
        """Generate Excel report based on type"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()[:31]  # Excel sheet name max 31 chars
        
        # Add header
        self._add_header(ws, report_type, filters)
        
        # Map report types to generators
        generators = {
            'financial_summary': self._create_financial_summary,
            'payroll_summary': self._create_payroll_summary,
            'inventory_status': self._create_inventory_status,
            'booking_summary': self._create_booking_summary,
            'employee_attendance': self._create_attendance_report,
            'revenue_analysis': self._create_revenue_analysis,
            'stock_movement': self._create_stock_movement,
            'occupancy_rate': self._create_occupancy_rate,
            'occupancy': self._create_occupancy_report,
            'daily_sales': self._create_daily_sales_report,
            'housekeeping': self._create_housekeeping_report,
            'maintenance': self._create_maintenance_report,
            'restaurant_sales': self._create_restaurant_sales_report,
            'bar_sales': self._create_bar_sales_report,
            'arrivals_departures': self._create_arrivals_departures_report,
            'room_supplies': self._create_room_supplies_report,
            'manager_duty': self._create_manager_duty_report,
            'reservation': self._create_reservation_report,
            'expense': self._create_expense_report,
            'branch_performance': self._create_branch_performance_report,
            'staff_overview': self._create_staff_overview_report,
            'compliance': self._create_compliance_report,
            'branch_comparison': self._create_branch_comparison_report,
        }
        
        generator = generators.get(report_type, self._create_generic_report)
        generator(ws, data)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
    
    def _add_header(self, ws, report_type, filters):
        """Add report header"""
        ws['A1'] = 'Famous Gate Hotel'
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A2'] = report_type.replace('_', ' ').title()
        ws['A2'].font = Font(bold=True, size=14)
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        
        if filters:
            filter_text = self._format_filters(filters)
            ws['A4'] = f"Filters: {filter_text}"
        
        # Add spacing
        ws.append([])
    
    def _format_filters(self, filters):
        """Format filters for display"""
        filter_parts = []
        if 'startDate' in filters:
            filter_parts.append(f"From {filters['startDate']}")
        if 'endDate' in filters:
            filter_parts.append(f"To {filters['endDate']}")
        if 'branch' in filters:
            filter_parts.append(f"Branch: {filters['branch']}")
        return ", ".join(filter_parts) if filter_parts else "None"
    
    def _create_financial_summary(self, ws, data):
        """Create financial summary in Excel"""
        # Find starting row
        start_row = 6
        
        # Add section header
        ws.cell(row=start_row, column=1, value="Financial Overview")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        # Add headers
        headers = ['Metric', 'Amount (KES)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Add data
        metrics = [
            ('Total Revenue', data.get('totalRevenue', 0)),
            ('Total Expenses', data.get('totalExpenses', 0)),
            ('Net Profit', data.get('netProfit', 0)),
            ('Operating Margin (%)', data.get('operatingMargin', 0))
        ]
        
        row = start_row + 2
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            if 'Margin' in metric:
                ws.cell(row=row, column=2, value=f"{value:.2f}%").border = self.border
            else:
                ws.cell(row=row, column=2, value=value).border = self.border
                ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1
        
        # Add revenue breakdown
        if 'revenueBreakdown' in data:
            row += 2
            ws.cell(row=row, column=1, value="Revenue Breakdown")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            
            row += 1
            for col, header in enumerate(['Source', 'Amount'], 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.border
            
            row += 1
            for source, amount in data['revenueBreakdown'].items():
                ws.cell(row=row, column=1, value=source.replace('_', ' ').title()).border = self.border
                ws.cell(row=row, column=2, value=amount).border = self.border
                ws.cell(row=row, column=2).number_format = '#,##0.00'
                row += 1
    
    def _create_payroll_summary(self, ws, data):
        """Create payroll summary in Excel"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Payroll Summary")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Employee ID', 'Name', 'Position', 'Gross Pay', 'Deductions', 'Net Pay']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        employees = data.get('employees', [])
        row = start_row + 2
        
        for emp in employees:
            ws.cell(row=row, column=1, value=emp.get('employeeId', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=emp.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=emp.get('position', 'N/A')).border = self.border
            ws.cell(row=row, column=4, value=emp.get('grossPay', 0)).border = self.border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=emp.get('deductions', 0)).border = self.border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=emp.get('netPay', 0)).border = self.border
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            row += 1
        
        # Add totals
        total_gross = sum(emp.get('grossPay', 0) for emp in employees)
        total_deductions = sum(emp.get('deductions', 0) for emp in employees)
        total_net = sum(emp.get('netPay', 0) for emp in employees)
        
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=3, value="").border = self.border
        ws.cell(row=row, column=4, value=total_gross).border = self.border
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_deductions).border = self.border
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_net).border = self.border
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=6).font = Font(bold=True)
    
    def _create_inventory_status(self, ws, data):
        """Create inventory status in Excel"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Inventory Status")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Item Code', 'Item Name', 'Category', 'Current Stock', 'Min Stock', 'Unit', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        items = data.get('items', [])
        row = start_row + 2
        
        for item in items:
            current_stock = item.get('currentStock', 0)
            min_stock = item.get('minStock', 0)
            status = 'Low Stock' if current_stock < min_stock else 'OK'
            
            ws.cell(row=row, column=1, value=item.get('itemCode', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=item.get('itemName', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=item.get('category', 'N/A')).border = self.border
            ws.cell(row=row, column=4, value=current_stock).border = self.border
            ws.cell(row=row, column=5, value=min_stock).border = self.border
            ws.cell(row=row, column=6, value=item.get('unit', 'pcs')).border = self.border
            
            status_cell = ws.cell(row=row, column=7, value=status)
            status_cell.border = self.border
            if status == 'Low Stock':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
    
    def _create_booking_summary(self, ws, data):
        """Create booking summary in Excel"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Booking Summary")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Booking ID', 'Guest Name', 'Room', 'Check-In', 'Check-Out', 'Nights', 'Amount', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        bookings = data.get('bookings', [])
        row = start_row + 2
        
        for booking in bookings:
            ws.cell(row=row, column=1, value=booking.get('bookingId', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=booking.get('guestName', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=booking.get('roomNumber', 'N/A')).border = self.border
            ws.cell(row=row, column=4, value=booking.get('checkIn', 'N/A')).border = self.border
            ws.cell(row=row, column=5, value=booking.get('checkOut', 'N/A')).border = self.border
            ws.cell(row=row, column=6, value=booking.get('nights', 0)).border = self.border
            ws.cell(row=row, column=7, value=booking.get('totalAmount', 0)).border = self.border
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=booking.get('status', 'N/A')).border = self.border
            row += 1
    
    def _create_attendance_report(self, ws, data):
        """Create attendance report in Excel"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Employee Attendance")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Employee ID', 'Name', 'Department', 'Date', 'Check-In', 'Check-Out', 'Hours']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        attendance = data.get('attendance', [])
        row = start_row + 2
        
        for record in attendance:
            ws.cell(row=row, column=1, value=record.get('employeeId', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=record.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=record.get('department', 'N/A')).border = self.border
            ws.cell(row=row, column=4, value=record.get('date', 'N/A')).border = self.border
            ws.cell(row=row, column=5, value=record.get('checkIn', 'N/A')).border = self.border
            ws.cell(row=row, column=6, value=record.get('checkOut', 'N/A')).border = self.border
            ws.cell(row=row, column=7, value=record.get('hours', 'N/A')).border = self.border
            row += 1
    
    def _create_revenue_analysis(self, ws, data):
        """Create revenue analysis report"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Revenue Analysis")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Date', 'Rooms', 'Restaurant', 'Bar', 'Other', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        revenue_data = data.get('dailyRevenue', [])
        row = start_row + 2
        
        for day in revenue_data:
            ws.cell(row=row, column=1, value=day.get('date', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=day.get('rooms', 0)).border = self.border
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=day.get('restaurant', 0)).border = self.border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=day.get('bar', 0)).border = self.border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=day.get('other', 0)).border = self.border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=day.get('total', 0)).border = self.border
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            row += 1
    
    def _create_stock_movement(self, ws, data):
        """Create stock movement report"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Stock Movement")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Date', 'Item Code', 'Item Name', 'Type', 'Quantity', 'From', 'To', 'Reference']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        movements = data.get('movements', [])
        row = start_row + 2
        
        for movement in movements:
            ws.cell(row=row, column=1, value=movement.get('date', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=movement.get('itemCode', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=movement.get('itemName', 'N/A')).border = self.border
            ws.cell(row=row, column=4, value=movement.get('type', 'N/A')).border = self.border
            ws.cell(row=row, column=5, value=movement.get('quantity', 0)).border = self.border
            ws.cell(row=row, column=6, value=movement.get('from', 'N/A')).border = self.border
            ws.cell(row=row, column=7, value=movement.get('to', 'N/A')).border = self.border
            ws.cell(row=row, column=8, value=movement.get('reference', 'N/A')).border = self.border
            row += 1
    
    def _create_occupancy_rate(self, ws, data):
        """Create occupancy rate report"""
        start_row = 6
        
        ws.cell(row=start_row, column=1, value="Room Occupancy Rate")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        headers = ['Date', 'Total Rooms', 'Occupied', 'Vacant', 'Occupancy %', 'Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        occupancy_data = data.get('dailyOccupancy', [])
        row = start_row + 2
        
        for day in occupancy_data:
            ws.cell(row=row, column=1, value=day.get('date', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=day.get('totalRooms', 0)).border = self.border
            ws.cell(row=row, column=3, value=day.get('occupied', 0)).border = self.border
            ws.cell(row=row, column=4, value=day.get('vacant', 0)).border = self.border
            ws.cell(row=row, column=5, value=f"{day.get('occupancyRate', 0):.2f}%").border = self.border
            ws.cell(row=row, column=6, value=day.get('revenue', 0)).border = self.border
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            row += 1
    
    def _create_generic_report(self, ws, data):
        """Create generic report for unknown types"""
        ws['A6'] = "Report Data"
        ws['A6'].font = Font(bold=True, size=12)
        row = 7
        for key, value in data.items():
            if not isinstance(value, (list, dict)):
                ws.cell(row=row, column=1, value=str(key)).border = self.border
                ws.cell(row=row, column=2, value=str(value)).border = self.border
                row += 1

    def _create_occupancy_report(self, ws, data):
        """Create occupancy report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Occupancy Summary").font = Font(bold=True, size=12)
        
        # Summary metrics
        metrics = [
            ('Occupancy Rate', f"{data.get('occupancy_rate', 0):.1f}%"),
            ('Total Rooms', data.get('total_rooms', 0)),
            ('Occupied Rooms', data.get('occupied_rooms', 0)),
            ('Available Rooms', data.get('available_rooms', 0)),
            ('ADR', f"KES {data.get('adr', 0):,.2f}"),
            ('RevPAR', f"KES {data.get('revpar', 0):,.2f}"),
        ]
        
        row = start_row + 1
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        # Room types
        row += 1
        ws.cell(row=row, column=1, value="By Room Type").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Room Type', 'Total', 'Occupied', 'Occupancy %', 'Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for rt in data.get('room_types', []):
            ws.cell(row=row, column=1, value=rt.get('type', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=rt.get('total', 0)).border = self.border
            ws.cell(row=row, column=3, value=rt.get('occupied', 0)).border = self.border
            ws.cell(row=row, column=4, value=f"{rt.get('occupancy_pct', 0):.1f}%").border = self.border
            ws.cell(row=row, column=5, value=rt.get('revenue', 0)).border = self.border
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            row += 1

    def _create_daily_sales_report(self, ws, data):
        """Create daily sales report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Sales Summary").font = Font(bold=True, size=12)
        
        metrics = [
            ('Total Revenue', f"KES {data.get('total_revenue', 0):,.2f}"),
            ('Total Transactions', data.get('total_transactions', 0)),
            ('Average Transaction', f"KES {data.get('avg_transaction', 0):,.2f}"),
            ('Cash Sales', f"KES {data.get('cash_sales', 0):,.2f}"),
            ('Card Sales', f"KES {data.get('card_sales', 0):,.2f}"),
            ('M-Pesa Sales', f"KES {data.get('mpesa_sales', 0):,.2f}"),
        ]
        
        row = start_row + 1
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        # Categories
        row += 1
        ws.cell(row=row, column=1, value="Sales by Category").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Category', 'Quantity', 'Total Sales']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for cat in data.get('categories', []):
            ws.cell(row=row, column=1, value=cat.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=cat.get('quantity', 0)).border = self.border
            ws.cell(row=row, column=3, value=cat.get('total', 0)).border = self.border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

    def _create_housekeeping_report(self, ws, data):
        """Create housekeeping report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Housekeeping Summary").font = Font(bold=True, size=12)
        
        metrics = [
            ('Rooms Cleaned', data.get('rooms_cleaned', 0)),
            ('Average Time (mins)', f"{data.get('avg_time', 0):.1f}"),
            ('Inspection Pass Rate', f"{data.get('pass_rate', 0):.1f}%"),
            ('Complaints', data.get('complaints', 0)),
        ]
        
        row = start_row + 1
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        # Staff performance
        row += 1
        ws.cell(row=row, column=1, value="Staff Performance").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Staff Name', 'Rooms Cleaned', 'Avg Time', 'Pass Rate', 'Rating']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for staff in data.get('staff', []):
            ws.cell(row=row, column=1, value=staff.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=staff.get('rooms', 0)).border = self.border
            ws.cell(row=row, column=3, value=f"{staff.get('avg_time', 0):.1f}").border = self.border
            ws.cell(row=row, column=4, value=f"{staff.get('pass_rate', 0):.1f}%").border = self.border
            ws.cell(row=row, column=5, value=f"{staff.get('rating', 0):.1f}/5").border = self.border
            row += 1

    def _create_maintenance_report(self, ws, data):
        """Create maintenance report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Maintenance Summary").font = Font(bold=True, size=12)
        
        metrics = [
            ('Total Requests', data.get('total_requests', 0)),
            ('Completed', data.get('completed', 0)),
            ('Pending', data.get('pending', 0)),
            ('In Progress', data.get('in_progress', 0)),
            ('Avg Resolution (hrs)', f"{data.get('avg_resolution', 0):.1f}"),
        ]
        
        row = start_row + 1
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        # Work orders
        row += 1
        ws.cell(row=row, column=1, value="Work Orders").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['ID', 'Location', 'Issue', 'Priority', 'Status', 'Assigned To']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for wo in data.get('work_orders', []):
            ws.cell(row=row, column=1, value=wo.get('id', 'N/A')[:8]).border = self.border
            ws.cell(row=row, column=2, value=wo.get('location', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=wo.get('issue', 'N/A')[:30]).border = self.border
            ws.cell(row=row, column=4, value=wo.get('priority', 'Normal')).border = self.border
            ws.cell(row=row, column=5, value=wo.get('status', 'Pending')).border = self.border
            ws.cell(row=row, column=6, value=wo.get('assigned_to', 'Unassigned')).border = self.border
            row += 1

    def _create_restaurant_sales_report(self, ws, data):
        """Create restaurant sales report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Restaurant Sales Summary").font = Font(bold=True, size=12)
        
        metrics = [
            ('Total Revenue', f"KES {data.get('total_revenue', 0):,.2f}"),
            ('Total Orders', data.get('total_orders', 0)),
            ('Average Order', f"KES {data.get('avg_order', 0):,.2f}"),
            ('Dine-In', data.get('dine_in', 0)),
            ('Room Service', data.get('room_service', 0)),
            ('Takeaway', data.get('takeaway', 0)),
        ]
        
        row = start_row + 1
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        # Top items
        row += 1
        ws.cell(row=row, column=1, value="Top Selling Items").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Item Name', 'Category', 'Quantity', 'Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for item in data.get('top_items', []):
            ws.cell(row=row, column=1, value=item.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=item.get('category', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=item.get('quantity', 0)).border = self.border
            ws.cell(row=row, column=4, value=item.get('revenue', 0)).border = self.border
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            row += 1

    def _create_bar_sales_report(self, ws, data):
        """Create bar sales report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Bar Sales Summary").font = Font(bold=True, size=12)
        
        metrics = [
            ('Total Revenue', f"KES {data.get('total_revenue', 0):,.2f}"),
            ('Total Orders', data.get('total_orders', 0)),
            ('Average Order', f"KES {data.get('avg_order', 0):,.2f}"),
        ]
        
        row = start_row + 1
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        # Top items
        row += 1
        ws.cell(row=row, column=1, value="Top Selling Drinks").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Item Name', 'Quantity', 'Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for item in data.get('top_items', []):
            ws.cell(row=row, column=1, value=item.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=item.get('quantity', 0)).border = self.border
            ws.cell(row=row, column=3, value=item.get('revenue', 0)).border = self.border
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

    def _create_arrivals_departures_report(self, ws, data):
        """Create arrivals and departures report"""
        start_row = 6
        
        # Arrivals
        ws.cell(row=start_row, column=1, value="Expected Arrivals").font = Font(bold=True, size=12)
        row = start_row + 1
        
        headers = ['Guest Name', 'Room', 'Expected Time', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        arrivals = data.get('arrivals', [])
        if not arrivals:
            ws.cell(row=row, column=1, value="No arrivals scheduled").border = self.border
            row += 1
        else:
            for arr in arrivals:
                ws.cell(row=row, column=1, value=arr.get('guest_name', 'N/A')).border = self.border
                ws.cell(row=row, column=2, value=arr.get('room', 'TBA')).border = self.border
                ws.cell(row=row, column=3, value=arr.get('time', 'N/A')).border = self.border
                ws.cell(row=row, column=4, value=arr.get('status', 'Expected')).border = self.border
                row += 1
        
        # Departures
        row += 1
        ws.cell(row=row, column=1, value="Expected Departures").font = Font(bold=True, size=12)
        row += 1
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        departures = data.get('departures', [])
        if not departures:
            ws.cell(row=row, column=1, value="No departures scheduled").border = self.border
        else:
            for dep in departures:
                ws.cell(row=row, column=1, value=dep.get('guest_name', 'N/A')).border = self.border
                ws.cell(row=row, column=2, value=dep.get('room', 'N/A')).border = self.border
                ws.cell(row=row, column=3, value=dep.get('time', 'N/A')).border = self.border
                ws.cell(row=row, column=4, value=dep.get('status', 'Expected')).border = self.border
                row += 1

    def _create_room_supplies_report(self, ws, data):
        """Create room supplies report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Room Supplies Inventory").font = Font(bold=True, size=12)
        row = start_row + 1
        
        headers = ['Item Name', 'Category', 'Current Stock', 'Unit', 'Reorder Level', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for item in data.get('supplies', []):
            qty = item.get('quantity', 0)
            reorder = item.get('reorder_level', 0)
            status = 'LOW' if qty < reorder else 'OK'
            
            ws.cell(row=row, column=1, value=item.get('name', 'N/A')).border = self.border
            ws.cell(row=row, column=2, value=item.get('category', 'N/A')).border = self.border
            ws.cell(row=row, column=3, value=qty).border = self.border
            ws.cell(row=row, column=4, value=item.get('unit', 'pcs')).border = self.border
            ws.cell(row=row, column=5, value=reorder).border = self.border
            status_cell = ws.cell(row=row, column=6, value=status)
            status_cell.border = self.border
            if status == 'LOW':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row += 1

    def _create_manager_duty_report(self, ws, data):
        """Create manager on duty report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Manager on Duty Report").font = Font(bold=True, size=12)
        
        info = [
            ('Manager', data.get('manager_name', 'N/A')),
            ('Date', data.get('date', 'N/A')),
            ('Shift', data.get('shift', 'Day')),
            ('Time', data.get('time', 'N/A')),
        ]
        
        row = start_row + 1
        for label, value in info:
            ws.cell(row=row, column=1, value=label).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="Operations Summary").font = Font(bold=True, size=12)
        row += 1
        
        ops = [
            ('Expected Arrivals', data.get('arrivals', 0)),
            ('Expected Departures', data.get('departures', 0)),
            ('Current Occupancy', f"{data.get('occupancy', 0)}%"),
            ('Out of Order Rooms', data.get('ooo_rooms', 0)),
        ]
        
        for label, value in ops:
            ws.cell(row=row, column=1, value=label).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1

    def _create_reservation_report(self, ws, data):
        """Create reservation confirmation report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Reservation Confirmation").font = Font(bold=True, size=12)
        
        guest_info = [
            ('Guest Name', data.get('guest_name', 'N/A')),
            ('Email', data.get('email', 'N/A')),
            ('Phone', data.get('phone', 'N/A')),
            ('ID Number', data.get('id_number', 'N/A')),
        ]
        
        row = start_row + 1
        for label, value in guest_info:
            ws.cell(row=row, column=1, value=label).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="Booking Details").font = Font(bold=True, size=12)
        row += 1
        
        booking_info = [
            ('Confirmation #', data.get('confirmation_number', 'N/A')),
            ('Room Type', data.get('room_type', 'N/A')),
            ('Room Number', data.get('room_number', 'TBA')),
            ('Check-In', data.get('check_in', 'N/A')),
            ('Check-Out', data.get('check_out', 'N/A')),
            ('Guests', data.get('guests', 1)),
        ]
        
        for label, value in booking_info:
            ws.cell(row=row, column=1, value=label).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="Payment Summary").font = Font(bold=True, size=12)
        row += 1
        
        payment_info = [
            ('Room Rate (per night)', f"KES {data.get('room_rate', 0):,.2f}"),
            ('Number of Nights', data.get('nights', 1)),
            ('Subtotal', f"KES {data.get('subtotal', 0):,.2f}"),
            ('Taxes & Fees', f"KES {data.get('taxes', 0):,.2f}"),
            ('TOTAL', f"KES {data.get('total', 0):,.2f}"),
            ('Payment Status', data.get('payment_status', 'Pending')),
        ]
        
        for label, value in payment_info:
            ws.cell(row=row, column=1, value=label).border = self.border
            cell = ws.cell(row=row, column=2, value=value)
            cell.border = self.border
            if label == 'TOTAL':
                cell.font = Font(bold=True)
            row += 1

    def _create_expense_report(self, ws, data):
        """Create expense report"""
        start_row = 6
        ws.cell(row=start_row, column=1, value="Expense Summary").font = Font(bold=True, size=12)
        
        ws.cell(row=start_row + 1, column=1, value="Total Expenses").border = self.border
        ws.cell(row=start_row + 1, column=2, value=f"KES {data.get('total_expenses', 0):,.2f}").border = self.border
        
        row = start_row + 3
        ws.cell(row=row, column=1, value="Expenses by Category").font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Category', 'Amount', '% of Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        for cat in data.get('by_category', []):
            ws.cell(row=row, column=1, value=cat.get('category', 'Other')).border = self.border
            ws.cell(row=row, column=2, value=cat.get('amount', 0)).border = self.border
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=f"{cat.get('pct', 0):.1f}%").border = self.border
            row += 1

    def _create_branch_performance_report(self, ws, data):
        """Create comprehensive branch performance report in Excel"""
        # Define colors for styling
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        start_row = 6
        
        # Executive Summary Section
        ws.cell(row=start_row, column=1, value="EXECUTIVE SUMMARY")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        ws.cell(row=start_row, column=1).fill = green_fill
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        
        row = start_row + 2
        summary_data = [
            ('Total Revenue', f"KES {data.get('total_revenue', 0):,.2f}"),
            ('Total Orders', f"{data.get('total_orders', 0):,}"),
            ('Average Satisfaction', f"{data.get('avg_satisfaction', 0):.1f}/5.0"),
            ('Best Performer', data.get('best_performer', 'N/A')),
            ('Needs Attention', data.get('worst_performer', 'N/A')),
        ]
        
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1
        
        row += 2
        
        # Branch Performance Comparison
        ws.cell(row=row, column=1, value="BRANCH PERFORMANCE COMPARISON")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = blue_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
        
        # Headers
        headers = ['Branch', 'Revenue', 'Change %', 'Orders', 'AOV', 'Rating', 'Efficiency', 'Target %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = yellow_fill
            cell.font = Font(bold=True)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Branch data
        branches = data.get('branches', [])
        for branch in branches:
            revenue_change = branch.get('revenue_change', 0)
            change_str = f"+{revenue_change:.1f}%" if revenue_change >= 0 else f"{revenue_change:.1f}%"
            target = branch.get('target_achievement', 0)
            
            ws.cell(row=row, column=1, value=branch.get('branch_name', 'Unknown')).border = self.border
            
            rev_cell = ws.cell(row=row, column=2, value=branch.get('revenue', 0))
            rev_cell.border = self.border
            rev_cell.number_format = '#,##0.00'
            
            change_cell = ws.cell(row=row, column=3, value=change_str)
            change_cell.border = self.border
            if revenue_change >= 0:
                change_cell.font = Font(color="006400")  # Green
            else:
                change_cell.font = Font(color="8B0000")  # Red
            
            ws.cell(row=row, column=4, value=branch.get('orders', 0)).border = self.border
            
            aov_cell = ws.cell(row=row, column=5, value=branch.get('avg_order_value', 0))
            aov_cell.border = self.border
            aov_cell.number_format = '#,##0.00'
            
            ws.cell(row=row, column=6, value=f"{branch.get('customer_satisfaction', 0):.1f}").border = self.border
            ws.cell(row=row, column=7, value=f"{branch.get('staff_efficiency', 0)}%").border = self.border
            
            target_cell = ws.cell(row=row, column=8, value=f"{target}%")
            target_cell.border = self.border
            if target >= 100:
                target_cell.fill = green_fill
            elif target >= 80:
                target_cell.fill = yellow_fill
            else:
                target_cell.fill = red_fill
            
            row += 1
        
        row += 2
        
        # Performance Rating Scale
        ws.cell(row=row, column=1, value="PERFORMANCE RATING SCALE")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = blue_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        ratings = [
            ('5 - Outstanding', 'Exceeds all targets', '00B050'),
            ('4 - Exceeds Expectations', 'Consistently exceeds targets', '92D050'),
            ('3 - Meets Expectations', 'Meets all required targets', 'FFFF00'),
            ('2 - Needs Improvement', 'Below target', 'FFC000'),
            ('1 - Unsatisfactory', 'Significantly below expectations', 'FF0000'),
        ]
        
        for rating, desc, color in ratings:
            cell = ws.cell(row=row, column=1, value=rating)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF" if color in ['00B050', 'FF0000'] else "000000")
            cell.border = self.border
            ws.cell(row=row, column=2, value=desc).border = self.border
            row += 1
        
        row += 2
        
        # Individual Branch Details
        for branch in branches:
            ws.cell(row=row, column=1, value=f"{branch.get('branch_name', 'Unknown')} - DETAILED METRICS")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = green_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
            
            # KPI headers
            kpi_headers = ['KPI', 'Value', 'Target', 'Status']
            for col, header in enumerate(kpi_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = gray_fill
                cell.font = Font(bold=True)
                cell.border = self.border
            row += 1
            
            # KPI data
            target_ach = branch.get('target_achievement', 100)
            revenue = branch.get('revenue', 0)
            target_revenue = revenue / (target_ach / 100) if target_ach > 0 else 0
            
            kpis = [
                ('Revenue', f"KES {revenue:,.2f}", f"KES {target_revenue:,.2f}", 
                 'On Track' if target_ach >= 80 else 'Below Target'),
                ('Orders', f"{branch.get('orders', 0):,}", '-', 'Active'),
                ('Customer Satisfaction', f"{branch.get('customer_satisfaction', 0):.1f}/5.0", '4.0/5.0',
                 'Good' if branch.get('customer_satisfaction', 0) >= 4.0 else 'Needs Improvement'),
                ('Staff Efficiency', f"{branch.get('staff_efficiency', 0)}%", '85%',
                 'Good' if branch.get('staff_efficiency', 0) >= 85 else 'Needs Improvement'),
                ('Inventory Turnover', f"{branch.get('inventory_turnover', 0)}x", '3x',
                 'Good' if branch.get('inventory_turnover', 0) >= 3 else 'Low'),
            ]
            
            for kpi, value, target, status in kpis:
                ws.cell(row=row, column=1, value=kpi).border = self.border
                ws.cell(row=row, column=2, value=value).border = self.border
                ws.cell(row=row, column=3, value=target).border = self.border
                status_cell = ws.cell(row=row, column=4, value=status)
                status_cell.border = self.border
                if 'Good' in status or 'On Track' in status or 'Active' in status:
                    status_cell.fill = green_fill
                else:
                    status_cell.fill = yellow_fill
                row += 1
            
            # Top Products
            top_products = branch.get('top_products', [])
            if top_products:
                row += 1
                ws.cell(row=row, column=1, value="Top Products").font = Font(bold=True)
                ws.cell(row=row, column=1).fill = yellow_fill
                ws.cell(row=row, column=2, value="Units Sold").font = Font(bold=True)
                ws.cell(row=row, column=2).fill = yellow_fill
                row += 1
                
                for prod in top_products[:5]:
                    ws.cell(row=row, column=1, value=prod.get('name', 'Unknown')).border = self.border
                    ws.cell(row=row, column=2, value=prod.get('sales', 0)).border = self.border
                    row += 1
            
            row += 2

    def _create_staff_overview_report(self, ws, data: Dict):
        """Create Staff Overview Excel report"""
        ws.title = "Staff Overview"
        
        # Header styling
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        yellow_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        
        row = 1
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "FAMOUS GATE HOTEL - STAFF OVERVIEW REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        row = 3
        
        # Executive Summary
        ws.cell(row=row, column=1, value="EXECUTIVE SUMMARY").font = Font(bold=True, size=12)
        row += 1
        
        summary_headers = ['Total Staff', 'Active', 'On Leave', 'Avg Performance', 'Avg Attendance']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = self.border
        row += 1
        
        summary_values = [
            data.get('total_staff', 0),
            data.get('active_staff', 0),
            data.get('on_leave', 0),
            f"{data.get('avg_performance', 0):.1f}%",
            f"{data.get('avg_attendance', 0):.1f}%"
        ]
        for col, value in enumerate(summary_values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        row += 3
        
        # Branch Staff Summary
        branch_summaries = data.get('branch_summaries', [])
        if branch_summaries:
            ws.cell(row=row, column=1, value="BRANCH STAFF SUMMARY").font = Font(bold=True, size=12)
            row += 1
            
            branch_headers = ['Branch', 'Total Staff', 'Active', 'On Leave', 'Avg Performance', 'Avg Attendance']
            for col, header in enumerate(branch_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = green_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = self.border
            row += 1
            
            for branch in branch_summaries:
                ws.cell(row=row, column=1, value=branch.get('branch_name', 'N/A')).border = self.border
                ws.cell(row=row, column=2, value=branch.get('total_staff', 0)).border = self.border
                ws.cell(row=row, column=3, value=branch.get('active', 0)).border = self.border
                ws.cell(row=row, column=4, value=branch.get('on_leave', 0)).border = self.border
                
                perf = branch.get('avg_performance', 0)
                perf_cell = ws.cell(row=row, column=5, value=f"{perf:.1f}%")
                perf_cell.border = self.border
                if perf >= 90:
                    perf_cell.fill = PatternFill(start_color="D1FAE5", fill_type="solid")
                elif perf < 70:
                    perf_cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
                
                att = branch.get('avg_attendance', 0)
                att_cell = ws.cell(row=row, column=6, value=f"{att:.1f}%")
                att_cell.border = self.border
                row += 1
            row += 2
        
        # Top Performers
        top_performers = data.get('top_performers', [])
        if top_performers:
            ws.cell(row=row, column=1, value="TOP PERFORMERS").font = Font(bold=True, size=12)
            row += 1
            
            perf_headers = ['Name', 'Branch', 'Role', 'Performance', 'Attendance', 'Rating']
            for col, header in enumerate(perf_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = green_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = self.border
            row += 1
            
            for staff in top_performers[:10]:
                ws.cell(row=row, column=1, value=staff.get('name', 'N/A')).border = self.border
                ws.cell(row=row, column=2, value=staff.get('branch_name', 'N/A')).border = self.border
                ws.cell(row=row, column=3, value=staff.get('role', 'N/A')).border = self.border
                ws.cell(row=row, column=4, value=f"{staff.get('performance_score', 0):.1f}%").border = self.border
                ws.cell(row=row, column=5, value=f"{staff.get('attendance_rate', 0):.1f}%").border = self.border
                ws.cell(row=row, column=6, value=f"{staff.get('customer_rating', 0):.1f}/5").border = self.border
                row += 1
            row += 2
        
        # Needs Attention
        needs_attention = data.get('needs_attention', [])
        if needs_attention:
            ws.cell(row=row, column=1, value="STAFF NEEDING ATTENTION").font = Font(bold=True, size=12)
            row += 1
            
            att_headers = ['Name', 'Branch', 'Role', 'Performance', 'Attendance', 'Issue']
            for col, header in enumerate(att_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = red_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = self.border
            row += 1
            
            for staff in needs_attention[:10]:
                issue = 'Low Performance' if staff.get('performance_score', 100) < 70 else 'Low Attendance'
                ws.cell(row=row, column=1, value=staff.get('name', 'N/A')).border = self.border
                ws.cell(row=row, column=2, value=staff.get('branch_name', 'N/A')).border = self.border
                ws.cell(row=row, column=3, value=staff.get('role', 'N/A')).border = self.border
                
                perf_cell = ws.cell(row=row, column=4, value=f"{staff.get('performance_score', 0):.1f}%")
                perf_cell.border = self.border
                if staff.get('performance_score', 100) < 70:
                    perf_cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
                
                att_cell = ws.cell(row=row, column=5, value=f"{staff.get('attendance_rate', 0):.1f}%")
                att_cell.border = self.border
                if staff.get('attendance_rate', 100) < 80:
                    att_cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
                
                ws.cell(row=row, column=6, value=issue).border = self.border
                row += 1

    def _create_compliance_report(self, ws, data: Dict):
        """Create Compliance Excel report"""
        ws.title = "Compliance Report"
        
        # Header styling
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        yellow_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        
        row = 1
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "FAMOUS GATE HOTEL - COMPLIANCE REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        row = 3
        
        # Compliance Overview
        ws.cell(row=row, column=1, value="COMPLIANCE OVERVIEW").font = Font(bold=True, size=12)
        row += 1
        
        overview_headers = ['Total Requirements', 'Compliant', 'Non-Compliant', 'Pending', 'Expired', 'Compliance Rate']
        for col, header in enumerate(overview_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = self.border
        row += 1
        
        compliance_rate = data.get('overall_compliance_rate', 0)
        overview_values = [
            data.get('total_requirements', 0),
            data.get('compliant', 0),
            data.get('non_compliant', 0),
            data.get('pending', 0),
            data.get('expired', 0),
            f"{compliance_rate:.1f}%"
        ]
        for col, value in enumerate(overview_values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            if col == 6:  # Compliance rate
                if compliance_rate >= 90:
                    cell.fill = green_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                elif compliance_rate >= 75:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
                    cell.font = Font(color="FFFFFF", bold=True)
        row += 3
        
        # Branch Compliance Summary
        branch_summaries = data.get('branch_summaries', [])
        if branch_summaries:
            ws.cell(row=row, column=1, value="BRANCH COMPLIANCE SUMMARY").font = Font(bold=True, size=12)
            row += 1
            
            branch_headers = ['Branch', 'Total', 'Compliant', 'Non-Compliant', 'Pending', 'Rate']
            for col, header in enumerate(branch_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = green_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = self.border
            row += 1
            
            for branch in branch_summaries:
                ws.cell(row=row, column=1, value=branch.get('branch_name', 'N/A')).border = self.border
                ws.cell(row=row, column=2, value=branch.get('total_requirements', 0)).border = self.border
                ws.cell(row=row, column=3, value=branch.get('compliant', 0)).border = self.border
                ws.cell(row=row, column=4, value=branch.get('non_compliant', 0)).border = self.border
                ws.cell(row=row, column=5, value=branch.get('pending', 0)).border = self.border
                
                rate = branch.get('compliance_rate', 0)
                rate_cell = ws.cell(row=row, column=6, value=f"{rate:.1f}%")
                rate_cell.border = self.border
                if rate >= 90:
                    rate_cell.fill = PatternFill(start_color="D1FAE5", fill_type="solid")
                elif rate < 75:
                    rate_cell.fill = PatternFill(start_color="FEE2E2", fill_type="solid")
                row += 1
            row += 2
        
        # Critical Issues
        critical_issues = data.get('critical_issues', [])
        if critical_issues:
            ws.cell(row=row, column=1, value="CRITICAL ISSUES").font = Font(bold=True, size=12)
            row += 1
            
            crit_headers = ['Requirement', 'Branch', 'Category', 'Status', 'Due Date']
            for col, header in enumerate(crit_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = red_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.border = self.border
            row += 1
            
            for item in critical_issues[:10]:
                ws.cell(row=row, column=1, value=item.get('requirement', 'N/A')[:50]).border = self.border
                ws.cell(row=row, column=2, value=item.get('branch_name', 'N/A')).border = self.border
                ws.cell(row=row, column=3, value=item.get('category', 'N/A')).border = self.border
                ws.cell(row=row, column=4, value=item.get('status', 'N/A').replace('_', ' ').title()).border = self.border
                due_date = item.get('due_date', '')
                ws.cell(row=row, column=5, value=due_date[:10] if due_date else 'N/A').border = self.border
                row += 1
            row += 2
        
        # Upcoming Deadlines
        upcoming = data.get('upcoming_deadlines', [])
        if upcoming:
            ws.cell(row=row, column=1, value="UPCOMING DEADLINES").font = Font(bold=True, size=12)
            row += 1
            
            up_headers = ['Requirement', 'Branch', 'Category', 'Due Date', 'Status']
            for col, header in enumerate(up_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
                cell.border = self.border
            row += 1
            
            for item in upcoming[:10]:
                ws.cell(row=row, column=1, value=item.get('requirement', 'N/A')[:50]).border = self.border
                ws.cell(row=row, column=2, value=item.get('branch_name', 'N/A')).border = self.border
                ws.cell(row=row, column=3, value=item.get('category', 'N/A')).border = self.border
                due_date = item.get('due_date', '')
                ws.cell(row=row, column=4, value=due_date[:10] if due_date else 'N/A').border = self.border
                ws.cell(row=row, column=5, value=item.get('status', 'N/A').replace('_', ' ').title()).border = self.border
                row += 1

    def _create_branch_comparison_report(self, ws, data):
        """Create branch comparison report worksheet"""
        # Extract data
        title = data.get('title', 'Branch Comparison Report')
        period = data.get('period', 'month')
        metric = data.get('metric', 'revenue')
        branches = data.get('branches', [])
        
        # Format labels
        period_label = {
            'week': 'Weekly',
            'month': 'Monthly',
            'quarter': 'Quarterly',
            'year': 'Yearly'
        }.get(period, 'Period')
        
        metric_label = {
            'revenue': 'Revenue (KES)',
            'occupancy': 'Occupancy Rate (%)',
            'staff_count': 'Staff Count'
        }.get(metric, metric.replace('_', ' ').capitalize())
        
        # Sort branches by metric value (descending)
        sorted_branches = sorted(branches, key=lambda x: x.get(metric, 0), reverse=True)
        
        # Add title
        ws.merge_cells('A1:D1')
        ws.cell(1, 1, title).font = Font(size=16, bold=True)
        
        # Add metadata
        report_date = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A2:D2')
        ws.cell(2, 1, report_date).font = Font(italic=True)
        
        ws.cell(3, 1, "Period:").font = Font(bold=True)
        ws.cell(3, 2, period_label)
        
        ws.cell(4, 1, "Metric:").font = Font(bold=True)
        ws.cell(4, 2, metric_label)
        
        # Add summary if branches exist
        row = 6
        if branches:
            best_branch = sorted_branches[0]['name']
            best_value = sorted_branches[0].get(metric, 0)
            
            # Format value based on metric type
            if metric == 'revenue':
                formatted_value = f"KES {best_value:,.2f}"
            elif metric == 'occupancy':
                formatted_value = f"{best_value}%"
            else:
                formatted_value = str(best_value)
            
            summary = f"Based on {metric_label} data for the selected {period}, {best_branch} is the top performing branch with a {metric_label} of {formatted_value}."
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row, 1, summary).font = Font(italic=True)
        
        # Add comparison table
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, "BRANCH COMPARISON").font = Font(bold=True, size=12)
        
        row += 1
        if branches:
            headers = ["Rank", "Branch", metric_label, "% of Top"]
            header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = header_font
                cell.fill = header_fill
            
            top_value = sorted_branches[0].get(metric, 1)  # Avoid division by zero
            
            for i, branch in enumerate(sorted_branches, 1):
                row += 1
                branch_name = branch.get('name', 'Unknown')
                value = branch.get(metric, 0)
                
                # Calculate percentage of top performer
                percent_of_top = (value / top_value * 100) if top_value else 0
                
                # Format value based on metric type
                if metric == 'revenue':
                    formatted_value = f"KES {value:,.2f}"
                elif metric == 'occupancy':
                    formatted_value = f"{value}%"
                else:
                    formatted_value = str(value)
                
                ws.cell(row, 1, i)
                ws.cell(row, 2, branch_name).font = Font(bold=True)
                ws.cell(row, 3, formatted_value)
                ws.cell(row, 4, f"{percent_of_top:.1f}%")
                
                # Add conditional formatting for visual indicators
                if i == 1:  # Top performer
                    ws.cell(row, 1).fill = PatternFill(start_color="D1FAE5", fill_type="solid")
                elif percent_of_top < 50:  # Low performers
                    ws.cell(row, 4).fill = PatternFill(start_color="FEE2E2", fill_type="solid")
        else:
            row += 1
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row, 1, "No branch comparison data available")
        
        # Add data analysis section
        row += 3
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, "DATA ANALYSIS").font = Font(bold=True, size=12)
        
        row += 1
        if branches and len(branches) > 1:
            # Calculate statistics
            values = [b.get(metric, 0) for b in branches]
            avg_value = sum(values) / len(values)
            min_value = min(values)
            max_value = max(values)
            range_value = max_value - min_value
            
            # Format based on metric type
            if metric == 'revenue':
                avg_formatted = f"KES {avg_value:,.2f}"
                min_formatted = f"KES {min_value:,.2f}"
                max_formatted = f"KES {max_value:,.2f}"
                range_formatted = f"KES {range_value:,.2f}"
            elif metric == 'occupancy':
                avg_formatted = f"{avg_value:.1f}%"
                min_formatted = f"{min_value}%"
                max_formatted = f"{max_value}%"
                range_formatted = f"{range_value}%"
            else:
                avg_formatted = f"{avg_value:.1f}"
                min_formatted = str(min_value)
                max_formatted = str(max_value)
                range_formatted = str(range_value)
            
            # Add statistics table
            headers = ["Statistic", "Value"]
            for col, header in enumerate(headers, 1):
                ws.cell(row, col, header).font = Font(bold=True)
                ws.cell(row, col).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            row += 1
            ws.cell(row, 1, "Average").font = Font(bold=True)
            ws.cell(row, 2, avg_formatted)
            
            row += 1
            ws.cell(row, 1, "Minimum").font = Font(bold=True)
            ws.cell(row, 2, min_formatted)
            
            row += 1
            ws.cell(row, 1, "Maximum").font = Font(bold=True)
            ws.cell(row, 2, max_formatted)
            
            row += 1
            ws.cell(row, 1, "Range").font = Font(bold=True)
            ws.cell(row, 2, range_formatted)
        else:
            row += 1
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row, 1, "Insufficient data for statistical analysis")
        
        # Add notes
        row += 3
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, "NOTES").font = Font(bold=True, size=12)
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, f"This report provides a comparison of branch performance based on {metric_label.lower()}.").font = Font(italic=True)
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, f"The data shown is for {period_label.lower()} period.").font = Font(italic=True)
        
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, "Rankings are based on current data and may change over time.").font = Font(italic=True)

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
