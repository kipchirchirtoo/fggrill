"""
Branch comparison Excel report generator function
"""

def _create_branch_comparison_report(self, ws, data):
    """Create branch comparison report worksheet"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Extract data
    title = data.get('title', 'Branch Comparison Report')
    period = data.get('period', 'month')
    metric = data.get('metric', 'revenue')
    branches = data.get('branches', [])
    
    # Format labels
    period_label = {
        'week': 'Weekly',
        'month': 'Monthly',
        'quarter': 'Quarterly',
        'year': 'Yearly'
    }.get(period, 'Period')
    
    metric_label = {
        'revenue': 'Revenue (KES)',
        'occupancy': 'Occupancy Rate (%)',
        'staff_count': 'Staff Count'
    }.get(metric, metric.replace('_', ' ').capitalize())
    
    # Sort branches by metric value (descending)
    sorted_branches = sorted(branches, key=lambda x: x.get(metric, 0), reverse=True)
    
    # Add title
    ws.merge_cells('A1:D1')
    ws.cell(1, 1, title).font = Font(size=16, bold=True)
    
    # Add metadata
    report_date = f"Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells('A2:D2')
    ws.cell(2, 1, report_date).font = Font(italic=True)
    
    ws.cell(3, 1, "Period:").font = Font(bold=True)
    ws.cell(3, 2, period_label)
    
    ws.cell(4, 1, "Metric:").font = Font(bold=True)
    ws.cell(4, 2, metric_label)
    
    # Add summary if branches exist
    row = 6
    if branches:
        best_branch = sorted_branches[0]['name']
        best_value = sorted_branches[0].get(metric, 0)
        
        # Format value based on metric type
        if metric == 'revenue':
            formatted_value = f"KES {best_value:,.2f}"
        elif metric == 'occupancy':
            formatted_value = f"{best_value}%"
        else:
            formatted_value = str(best_value)
        
        summary = f"Based on {metric_label} data for the selected {period}, {best_branch} is the top performing branch with a {metric_label} of {formatted_value}."
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, summary).font = Font(italic=True)
    
    # Add comparison table
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, "BRANCH COMPARISON").font = Font(bold=True, size=12)
    
    row += 1
    if branches:
        headers = ["Rank", "Branch", metric_label, "% of Top"]
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = header_font
            cell.fill = header_fill
        
        top_value = sorted_branches[0].get(metric, 1)  # Avoid division by zero
        
        for i, branch in enumerate(sorted_branches, 1):
            row += 1
            branch_name = branch.get('name', 'Unknown')
            value = branch.get(metric, 0)
            
            # Calculate percentage of top performer
            percent_of_top = (value / top_value * 100) if top_value else 0
            
            # Format value based on metric type
            if metric == 'revenue':
                formatted_value = f"KES {value:,.2f}"
            elif metric == 'occupancy':
                formatted_value = f"{value}%"
            else:
                formatted_value = str(value)
            
            ws.cell(row, 1, i)
            ws.cell(row, 2, branch_name).font = Font(bold=True)
            ws.cell(row, 3, formatted_value)
            ws.cell(row, 4, f"{percent_of_top:.1f}%")
            
            # Add conditional formatting for visual indicators
            if i == 1:  # Top performer
                ws.cell(row, 1).fill = PatternFill(start_color="D1FAE5", fill_type="solid")
            elif percent_of_top < 50:  # Low performers
                ws.cell(row, 4).fill = PatternFill(start_color="FEE2E2", fill_type="solid")
    else:
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, "No branch comparison data available")
    
    # Add data analysis section
    row += 3
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, "DATA ANALYSIS").font = Font(bold=True, size=12)
    
    row += 1
    if branches and len(branches) > 1:
        # Calculate statistics
        values = [b.get(metric, 0) for b in branches]
        avg_value = sum(values) / len(values)
        min_value = min(values)
        max_value = max(values)
        range_value = max_value - min_value
        
        # Format based on metric type
        if metric == 'revenue':
            avg_formatted = f"KES {avg_value:,.2f}"
            min_formatted = f"KES {min_value:,.2f}"
            max_formatted = f"KES {max_value:,.2f}"
            range_formatted = f"KES {range_value:,.2f}"
        elif metric == 'occupancy':
            avg_formatted = f"{avg_value:.1f}%"
            min_formatted = f"{min_value}%"
            max_formatted = f"{max_value}%"
            range_formatted = f"{range_value}%"
        else:
            avg_formatted = f"{avg_value:.1f}"
            min_formatted = str(min_value)
            max_formatted = str(max_value)
            range_formatted = str(range_value)
        
        # Add statistics table
        headers = ["Statistic", "Value"]
        for col, header in enumerate(headers, 1):
            ws.cell(row, col, header).font = Font(bold=True)
            ws.cell(row, col).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        row += 1
        ws.cell(row, 1, "Average").font = Font(bold=True)
        ws.cell(row, 2, avg_formatted)
        
        row += 1
        ws.cell(row, 1, "Minimum").font = Font(bold=True)
        ws.cell(row, 2, min_formatted)
        
        row += 1
        ws.cell(row, 1, "Maximum").font = Font(bold=True)
        ws.cell(row, 2, max_formatted)
        
        row += 1
        ws.cell(row, 1, "Range").font = Font(bold=True)
        ws.cell(row, 2, range_formatted)
    else:
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row, 1, "Insufficient data for statistical analysis")
    
    # Add notes
    row += 3
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, "NOTES").font = Font(bold=True, size=12)
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, f"This report provides a comparison of branch performance based on {metric_label.lower()}.").font = Font(italic=True)
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, f"The data shown is for {period_label.lower()} period.").font = Font(italic=True)
    
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row, 1, "Rankings are based on current data and may change over time.").font = Font(italic=True)
