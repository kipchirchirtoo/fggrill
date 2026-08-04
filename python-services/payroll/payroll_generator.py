# -*- coding: utf-8 -*-
"""
FAMOUSGATE HOTELS - Payroll XLSX + PDF Generator
"""
import io, os
from datetime import datetime

LOGO_PATH = os.path.join(os.path.dirname(__file__), '..', 'assets', 'fglogo.png')

THEME = {
    "dark":       "1A4731",
    "mid":        "2D6A4F",
    "accent":     "40916C",
    "light":      "D8F3DC",
    "light_gray": "F2F2F2",
    "draft_bg":   "FFF3CD",
    "draft_fg":   "856404",
    "border":     "B0B0B0",
}

# (header_label, width, data_key, align, number_format)
COLUMNS = [
    ("#",                      5,  "no",           "center", None),
    ("Emp ID",                13,  "emp_id",       "center", None),
    ("Staff Name",            30,  "name",         "left",   None),
    ("Phone Number",          20,  "phone",        "center", None),
    ("Role",                  24,  "role",         "left",   None),
    ("Branch",                16,  "branch",       "left",   None),
    ("Basic Salary\n(KES)",   16,  "basic_salary", "right",  "#,##0.00"),
    ("NSSF\n(KES)",           14,  "nssf",         "right",  "#,##0.00"),
    ("SHIF\n(KES)",           14,  "shif",         "right",  "#,##0.00"),
    ("Housing Levy\n(KES)",   14,  "housing_levy", "right",  "#,##0.00"),
    ("PAYE\n(KES)",           14,  "paye",         "right",  "#,##0.00"),
    ("Credit Bills\n(KES)",   14,  "credit_bills", "right",  "#,##0.00"),
    ("Advances\n(KES)",       14,  "advances",     "right",  "#,##0.00"),
    ("Loans\n(KES)",          14,  "loans",        "right",  "#,##0.00"),
    ("Total Deductions\n(KES)",16, "FORMULA_DEDUCTIONS", "right", "#,##0.00"),
    ("Net Pay\n(KES)",        16,  "FORMULA_NET_PAY",    "right", "#,##0.00"),
    ("Status",                12,  "STATUS",       "center", None),
]

DEDUCTION_COLS = [8, 9, 10, 11, 12, 13, 14]
BASIC_COL      = 7
TOTAL_DED_COL  = 15
NET_PAY_COL    = 16
STATUS_COL     = 17
TOTAL_COLS     = len(COLUMNS)


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX GENERATOR
#  Layout: Row 1 = column headers (FROZEN), rows 2..N = data, rows N+2.. = branding
# ══════════════════════════════════════════════════════════════════════════════
def generate_payroll_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    employees = data.get("employees", [])
    period    = data.get("period", "")
    branch    = data.get("branch", "All Branches")
    generated = data.get("generated", datetime.now().strftime("%d/%m/%Y, %H:%M:%S"))
    status    = data.get("status", "DRAFT")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {period}"

    def F(bold=False, size=11, color="000000", italic=False):
        return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    def BG(hex_color):
        return PatternFill("solid", fgColor=hex_color)
    def AL(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def thin():
        s = Side(style="thin", color=THEME["border"])
        return Border(left=s, right=s, top=s, bottom=s)
    def thick():
        s = Side(style="medium", color="555555")
        return Border(left=s, right=s, top=s, bottom=s)

    def merge_row(row, sc, ec, value, font, bg, alignment, border=None):
        ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
        c = ws.cell(row, sc)
        c.value = value; c.font = font; c.fill = BG(bg); c.alignment = alignment
        if border: c.border = border

    # Column widths
    for i, (_, width, *_r) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── ROW 1: Column headers (ONLY this row is frozen) ──────────────────────
    ws.row_dimensions[1].height = 44
    for ci, (header, *_r) in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci); c.value = header
        c.font = F(bold=True, size=10, color="FFFFFF")
        c.fill = BG(THEME["dark"])
        c.alignment = AL("center", "center", wrap=True)
        c.border = thin()

    # Freeze ONLY the column header row
    ws.freeze_panes = "A2"

    # ── DATA ROWS (start at row 2) ────────────────────────────────────────────
    DATA_START = 2
    for idx, emp in enumerate(employees):
        r = DATA_START + idx
        ws.row_dimensions[r].height = 22
        row_bg = THEME["light"] if idx % 2 == 0 else "FFFFFF"
        for ci, (_, _, data_key, align_h, num_fmt) in enumerate(COLUMNS, 1):
            c = ws.cell(r, ci); c.fill = BG(row_bg); c.border = thin()
            if data_key == "FORMULA_DEDUCTIONS":
                ded_refs = "+".join(f"{get_column_letter(dc)}{r}" for dc in DEDUCTION_COLS)
                c.value = f"={ded_refs}"; c.number_format = num_fmt
                c.alignment = AL(align_h, "center"); c.font = F(size=10)
            elif data_key == "FORMULA_NET_PAY":
                c.value = f"={get_column_letter(BASIC_COL)}{r}-{get_column_letter(TOTAL_DED_COL)}{r}"
                c.number_format = num_fmt; c.alignment = AL(align_h, "center")
                c.font = F(bold=True, size=10, color=THEME["dark"])
            elif data_key == "STATUS":
                c.value = status
                c.font = Font(name="Arial", bold=True, size=9, color=THEME["draft_fg"])
                c.fill = BG(THEME["draft_bg"]); c.alignment = AL("center", "center")
            else:
                c.value = emp.get(data_key, "")
                if num_fmt: c.number_format = num_fmt
                c.alignment = AL(align_h, "center")
                if data_key == "emp_id":
                    c.font = F(bold=True, size=10, color=THEME["mid"])
                else:
                    c.font = F(size=10)

    # ── TOTALS ROW ────────────────────────────────────────────────────────────
    TOTALS_ROW = DATA_START + len(employees)
    LAST_DATA  = TOTALS_ROW - 1
    ws.row_dimensions[TOTALS_ROW].height = 26
    ws.merge_cells(start_row=TOTALS_ROW, start_column=1, end_row=TOTALS_ROW, end_column=6)
    c = ws.cell(TOTALS_ROW, 1); c.value = "TOTALS"
    c.font = F(bold=True, size=12, color="FFFFFF"); c.fill = BG(THEME["accent"])
    c.alignment = AL("center", "center"); c.border = thick()
    SUM_COLS = [BASIC_COL] + DEDUCTION_COLS + [TOTAL_DED_COL, NET_PAY_COL]
    for ci in range(1, TOTAL_COLS + 1):
        c = ws.cell(TOTALS_ROW, ci); c.fill = BG(THEME["accent"]); c.border = thick()
        if ci in SUM_COLS:
            cl = get_column_letter(ci)
            c.value = f"=SUM({cl}{DATA_START}:{cl}{LAST_DATA})"
            c.number_format = "#,##0.00"
            c.font = F(bold=True, size=11, color="FFFFFF"); c.alignment = AL("right", "center")

    # ── BRANDING BLOCK (below data, scrolls with sheet) ──────────────────────
    B = TOTALS_ROW + 2  # branding start row

    # Summary header
    ws.row_dimensions[B].height = 22
    merge_row(B, 1, TOTAL_COLS, "PAYROLL SUMMARY",
              F(bold=True, size=11, color="FFFFFF"), THEME["accent"], AL("center", "center"))

    # Summary boxes
    total_basic = sum(e.get("basic_salary", 0) for e in employees)
    total_ded   = sum(e.get("nssf",0)+e.get("shif",0)+e.get("housing_levy",0)+
                      e.get("paye",0)+e.get("credit_bills",0)+e.get("advances",0)+e.get("loans",0)
                      for e in employees)
    total_net   = total_basic - total_ded
    summary_boxes = [
        ("Total Staff",              len(employees),          1,  4),
        ("Total Basic Salary (KES)", f"{total_basic:,.2f}",   5,  9),
        ("Total Deductions (KES)",   f"{total_ded:,.2f}",     10, 13),
        ("Total Net Pay (KES)",      f"{total_net:,.2f}",     14, TOTAL_COLS),
    ]
    ws.row_dimensions[B+1].height = 18; ws.row_dimensions[B+2].height = 26
    for label, value, sc, ec in summary_boxes:
        ws.merge_cells(start_row=B+1, start_column=sc, end_row=B+1, end_column=ec)
        ws.merge_cells(start_row=B+2, start_column=sc, end_row=B+2, end_column=ec)
        lc = ws.cell(B+1, sc); lc.value = label
        lc.font = Font(name="Arial", size=9, bold=True, color="555555")
        lc.fill = BG(THEME["light_gray"]); lc.alignment = AL("center","center"); lc.border = thin()
        vc = ws.cell(B+2, sc); vc.value = value
        vc.font = Font(name="Arial", size=13, bold=True, color=THEME["dark"])
        vc.fill = BG("FFFFFF"); vc.alignment = AL("center","center"); vc.border = thick()

    # Meta info
    M = B + 4
    label_font = Font(name="Arial", bold=True, size=10, color=THEME["dark"])
    value_font = Font(name="Arial", size=10, color="222222")
    meta_rows = [
        ("Period:",      period,    "Branch:",      branch),
        ("Generated:",   generated, "Status:",      status),
        ("Prepared by:", "_" * 33,  "Approved by:", "_" * 33),
    ]
    for offset, (l1, v1, l2, v2) in enumerate(meta_rows):
        r = M + offset
        ws.row_dimensions[r].height = 20
        for sc, ec, val, fnt in [(1,3,l1,label_font),(4,9,v1,value_font),(10,12,l2,label_font),(13,TOTAL_COLS,v2,value_font)]:
            ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=ec)
            c = ws.cell(r, sc)
            c.value = val; c.font = fnt
            c.fill = BG(THEME["light_gray"]) if val in (l1, l2) else BG("FFFFFF")
            c.alignment = AL("left", "center"); c.border = thin()

    # Company header
    H = M + len(meta_rows) + 2
    ws.row_dimensions[H].height = 45
    merge_row(H, 1, TOTAL_COLS, data.get("company_name", "FAMOUSGATE HOTELS"),
              F(bold=True, size=22, color="FFFFFF"), THEME["dark"], AL("center", "center"))
    ws.row_dimensions[H+1].height = 22
    contact = (f"{data.get('company_address','Bomet, Kenya')}   |   "
               f"{data.get('company_email','famousgateshotelsbmt@gmail.com')}"
               f"   |   Tel: {data.get('company_phone','0706 782 828')}")
    merge_row(H+1, 1, TOTAL_COLS, contact,
              F(size=10, italic=True, color="FFFFFF"), THEME["mid"], AL("center","center"))

    # Footer
    FOOTER_ROW = H + 3
    ws.row_dimensions[FOOTER_ROW].height = 18
    merge_row(FOOTER_ROW, 1, TOTAL_COLS,
              f"© 2026 {data.get('company_name','FAMOUSGATE HOTELS')} — Confidential Payroll Document",
              Font(name="Arial", size=9, italic=True, color="888888"),
              THEME["light_gray"], AL("center","center"))

    # Sheet settings
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
#  PDF GENERATOR  (landscape A4, logo, branch-aware)
# ══════════════════════════════════════════════════════════════════════════════
def generate_payroll_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable, Image)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    DARK_GREEN  = colors.HexColor("#1A4731")
    MID_GREEN   = colors.HexColor("#2D6A4F")
    ACCENT      = colors.HexColor("#40916C")
    LIGHT_GREEN = colors.HexColor("#D8F3DC")
    LIGHT_GRAY  = colors.HexColor("#F2F2F2")
    DRAFT_BG    = colors.HexColor("#FFF3CD")
    DRAFT_FG    = colors.HexColor("#856404")
    WHITE       = colors.white
    BORDER_CLR  = colors.HexColor("#B0B0B0")

    def S(name, **kw): return ParagraphStyle(name, **kw)
    STYLES = {
        "company":    S("company",    fontName="Helvetica-Bold",   fontSize=18, textColor=WHITE,      alignment=TA_CENTER),
        "contact":    S("contact",    fontName="Helvetica-Oblique",fontSize=8,  textColor=WHITE,      alignment=TA_CENTER),
        "title":      S("title",      fontName="Helvetica-Bold",   fontSize=14, textColor=DARK_GREEN, alignment=TA_CENTER, spaceBefore=4, spaceAfter=4),
        "branch_tag": S("branch_tag", fontName="Helvetica-Bold",   fontSize=11, textColor=WHITE,      alignment=TA_CENTER),
        "meta_label": S("meta_label", fontName="Helvetica-Bold",   fontSize=8,  textColor=DARK_GREEN),
        "meta_value": S("meta_value", fontName="Helvetica",        fontSize=8,  textColor=colors.black),
        "sum_lbl":    S("sum_lbl",    fontName="Helvetica-Bold",   fontSize=7,  textColor=colors.HexColor("#555555"), alignment=TA_CENTER),
        "sum_val":    S("sum_val",    fontName="Helvetica-Bold",   fontSize=12, textColor=DARK_GREEN, alignment=TA_CENTER),
        "col_hdr":    S("col_hdr",    fontName="Helvetica-Bold",   fontSize=6.5,textColor=WHITE,      alignment=TA_CENTER, leading=8),
        "cell":       S("cell",       fontName="Helvetica",        fontSize=7,  textColor=colors.black, leading=9),
        "cell_r":     S("cell_r",     fontName="Helvetica",        fontSize=7,  textColor=colors.black, alignment=TA_RIGHT, leading=9),
        "cell_c":     S("cell_c",     fontName="Helvetica",        fontSize=7,  textColor=colors.black, alignment=TA_CENTER, leading=9),
        "emp_id":     S("emp_id",     fontName="Helvetica-Bold",   fontSize=7,  textColor=MID_GREEN,  alignment=TA_CENTER, leading=9),
        "total_lbl":  S("total_lbl",  fontName="Helvetica-Bold",   fontSize=8,  textColor=WHITE,      alignment=TA_CENTER),
        "total_val":  S("total_val",  fontName="Helvetica-Bold",   fontSize=8,  textColor=WHITE,      alignment=TA_RIGHT, leading=9),
        "status":     S("status",     fontName="Helvetica-Bold",   fontSize=6.5,textColor=DRAFT_FG,   alignment=TA_CENTER),
        "sig":        S("sig",        fontName="Helvetica-Oblique",fontSize=8,  textColor=colors.HexColor("#444444")),
        "footer":     S("footer",     fontName="Helvetica-Oblique",fontSize=7,  textColor=colors.HexColor("#888888"), alignment=TA_CENTER),
        "net_pay":    S("net_pay",    fontName="Helvetica-Bold",   fontSize=7,  textColor=DARK_GREEN, alignment=TA_RIGHT, leading=9),
    }

    def P(text, style): return Paragraph(str(text), STYLES[style])
    def fmt(v):
        try: return f"{float(v):,.2f}"
        except: return "0.00"

    employees   = data.get("employees", [])
    period      = data.get("period", "")
    branch      = data.get("branch", "All Branches")
    generated   = data.get("generated", datetime.now().strftime("%d/%m/%Y, %H:%M:%S"))
    status      = data.get("status", "DRAFT")
    company     = data.get("company_name", "FAMOUSGATE HOTELS")
    address     = data.get("company_address", "Bomet, Kenya")
    email       = data.get("company_email", "famousgateshotelsbmt@gmail.com")
    phone       = data.get("company_phone", "0706 782 828")

    page_w, page_h = landscape(A4)
    margin = 12 * mm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=margin)
    usable_w = page_w - 2 * margin
    story = []

    # ── HEADER BANNER with logo ───────────────────────────────────────────────
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        try:
            logo_img = Image(LOGO_PATH, width=18*mm, height=18*mm)
            logo_cell = logo_img
        except Exception:
            logo_cell = P("FG", "company")

    logo_w  = 22 * mm
    text_w  = usable_w - logo_w

    hdr_inner = Table(
        [[logo_cell, P(company, "company")]],
        colWidths=[logo_w, text_w]
    )
    hdr_inner.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), DARK_GREEN),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",(0,0),(0,0), 4),
        ("TOPPADDING", (0,0),(-1,-1), 10),
        ("BOTTOMPADDING",(0,0),(-1,-1), 10),
    ]))

    contact_row = Table(
        [[P(f"{address}   |   {email}   |   Tel: {phone}", "contact")]],
        colWidths=[usable_w]
    )
    contact_row.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), MID_GREEN),
        ("TOPPADDING",(0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))

    story += [hdr_inner, contact_row, Spacer(1, 3*mm)]

    # ── TITLE + BRANCH TAG ────────────────────────────────────────────────────
    story.append(P("PAYROLL REPORT", "title"))

    # Branch tag — only show if not "All Branches"
    if branch and branch != "All Branches":
        branch_tbl = Table([[P(f"Branch: {branch}", "branch_tag")]],
                           colWidths=[usable_w])
        branch_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), ACCENT),
            ("TOPPADDING",(0,0),(-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("ROUNDEDCORNERS", [3]),
        ]))
        story.append(branch_tbl)
        story.append(Spacer(1, 2*mm))

    story.append(HRFlowable(width=usable_w, thickness=2, color=DARK_GREEN, spaceAfter=3*mm))

    # ── META ──────────────────────────────────────────────────────────────────
    hw = usable_w / 2
    meta = Table([
        [P("Period:",    "meta_label"), P(period,    "meta_value"), P("Branch:",  "meta_label"), P(branch,    "meta_value")],
        [P("Generated:","meta_label"), P(generated, "meta_value"), P("Status:",  "meta_label"), P(status,    "meta_value")],
    ], colWidths=[25*mm, hw-25*mm, 25*mm, hw-25*mm])
    meta.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,-1),LIGHT_GRAY),("BACKGROUND",(2,0),(2,-1),LIGHT_GRAY),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),4),("GRID",(0,0),(-1,-1),0.4,BORDER_CLR),
    ]))
    story += [meta, Spacer(1, 3*mm)]

    # ── SUMMARY BOXES ─────────────────────────────────────────────────────────
    total_basic = sum(e.get("basic_salary",0) for e in employees)
    total_ded   = sum(e.get("nssf",0)+e.get("shif",0)+e.get("housing_levy",0)+
                      e.get("paye",0)+e.get("credit_bills",0)+e.get("advances",0)+e.get("loans",0)
                      for e in employees)
    total_net   = total_basic - total_ded
    qw = usable_w / 4
    sum_tbl = Table([
        [P("Total Staff","sum_lbl"), P("Total Basic Salary (KES)","sum_lbl"),
         P("Total Deductions (KES)","sum_lbl"), P("Total Net Pay (KES)","sum_lbl")],
        [P(str(len(employees)),"sum_val"), P(fmt(total_basic),"sum_val"),
         P(fmt(total_ded),"sum_val"),       P(fmt(total_net),"sum_val")],
    ], colWidths=[qw,qw,qw,qw])
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),LIGHT_GRAY),("BACKGROUND",(0,1),(-1,1),WHITE),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("GRID",(0,0),(-1,-1),0.8,colors.HexColor("#555555")),
        ("LINEABOVE",(0,0),(-1,0),1.5,DARK_GREEN),("LINEBELOW",(0,1),(-1,1),1.5,DARK_GREEN),
    ]))
    story += [sum_tbl, Spacer(1, 4*mm)]

    # ── PAYROLL TABLE ─────────────────────────────────────────────────────────
    col_mm  = [7,14,30,20,24,16,18,13,13,14,14,13,13,13,18,18,12]
    col_pts = [c*mm for c in col_mm]
    headers = [
        P("#","col_hdr"), P("Emp ID","col_hdr"), P("Staff Name","col_hdr"),
        P("Phone Number","col_hdr"), P("Role","col_hdr"), P("Branch","col_hdr"),
        P("Basic Salary\n(KES)","col_hdr"), P("NSSF\n(KES)","col_hdr"),
        P("SHIF\n(KES)","col_hdr"), P("Housing\nLevy (KES)","col_hdr"),
        P("PAYE\n(KES)","col_hdr"), P("Credit\nBills (KES)","col_hdr"),
        P("Advances\n(KES)","col_hdr"), P("Loans\n(KES)","col_hdr"),
        P("Total\nDeductions\n(KES)","col_hdr"), P("Net Pay\n(KES)","col_hdr"),
        P("Status","col_hdr"),
    ]
    table_data = [headers]
    for emp in employees:
        ded = (emp.get("nssf",0)+emp.get("shif",0)+emp.get("housing_levy",0)+
               emp.get("paye",0)+emp.get("credit_bills",0)+emp.get("advances",0)+emp.get("loans",0))
        net = emp.get("basic_salary",0) - ded
        table_data.append([
            P(str(emp.get("no","")),          "cell_c"),
            P(emp.get("emp_id",""),            "emp_id"),
            P(emp.get("name",""),              "cell"),
            P(emp.get("phone",""),             "cell_c"),
            P(emp.get("role",""),              "cell"),
            P(emp.get("branch",""),            "cell"),
            P(fmt(emp.get("basic_salary",0)),  "cell_r"),
            P(fmt(emp.get("nssf",0)),          "cell_r"),
            P(fmt(emp.get("shif",0)),          "cell_r"),
            P(fmt(emp.get("housing_levy",0)),  "cell_r"),
            P(fmt(emp.get("paye",0)),          "cell_r"),
            P(fmt(emp.get("credit_bills",0)),  "cell_r"),
            P(fmt(emp.get("advances",0)),      "cell_r"),
            P(fmt(emp.get("loans",0)),         "cell_r"),
            P(fmt(ded),                        "cell_r"),
            P(fmt(net),                        "net_pay"),
            P(status,                          "status"),
        ])
    # Totals row
    table_data.append([
        P("TOTALS","total_lbl"),"","","","","",
        P(fmt(sum(e.get("basic_salary",0) for e in employees)),    "total_val"),
        P(fmt(sum(e.get("nssf",0)         for e in employees)),    "total_val"),
        P(fmt(sum(e.get("shif",0)         for e in employees)),    "total_val"),
        P(fmt(sum(e.get("housing_levy",0) for e in employees)),    "total_val"),
        P(fmt(sum(e.get("paye",0)         for e in employees)),    "total_val"),
        P(fmt(sum(e.get("credit_bills",0) for e in employees)),    "total_val"),
        P(fmt(sum(e.get("advances",0)     for e in employees)),    "total_val"),
        P(fmt(sum(e.get("loans",0)        for e in employees)),    "total_val"),
        P(fmt(total_ded),                                          "total_val"),
        P(fmt(total_net),                                          "total_val"),
        P("","total_val"),
    ])
    n_rows   = len(table_data)
    last_row = n_rows - 1
    tbl = Table(table_data, colWidths=col_pts, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),DARK_GREEN),
        ("ROWBACKGROUNDS",(0,1),(-1,last_row-1),[WHITE,LIGHT_GREEN]),
        ("BACKGROUND",(0,last_row),(-1,last_row),ACCENT),
        ("SPAN",(0,last_row),(5,last_row)),
        *[("BACKGROUND",(16,i),(16,i),DRAFT_BG) for i in range(1,last_row)],
        ("GRID",(0,0),(-1,-1),0.4,BORDER_CLR),
        ("LINEBELOW",(0,0),(-1,0),1.0,WHITE),
        ("LINEABOVE",(0,last_row),(-1,last_row),1.5,colors.HexColor("#555555")),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    story += [tbl, Spacer(1, 8*mm)]

    # ── SIGNATURES ────────────────────────────────────────────────────────────
    sig = Table([[P("Prepared by: "+"_"*38,"sig"), P("","sig"), P("Approved by: "+"_"*38,"sig")]],
                colWidths=[usable_w*0.44, usable_w*0.12, usable_w*0.44])
    sig.setStyle(TableStyle([("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    story += [sig, Spacer(1, 4*mm),
              HRFlowable(width=usable_w, thickness=0.5, color=BORDER_CLR),
              Spacer(1, 2*mm),
              P(f"© 2026 {company} — Confidential Payroll Document   |   Generated: {generated}", "footer")]

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════════════════════
#  GENERIC STATEMENT PDF  (same FamousGate branding as payroll, arbitrary table)
#  Used by the Branch Accountant "Staff Accounts" exports (credit bills, salary
#  advances, staff loans, combined overview). The caller supplies pre-formatted
#  cell strings so this stays fully generic.
#
#  data = {
#    "title":   "STAFF CREDIT BILLS",
#    "period":  "05 Jul 2026 - 04 Aug 2026",
#    "branch":  "Kyogong",
#    "generated": "...",              # optional
#    "columns": [ {"header": "Date", "align": "center", "weight": 1.2}, ... ],
#    "rows":    [ ["04/08/26", "John Doe", "1,200.00", ...], ... ],
#    "summary": [ {"label": "Records", "value": "12"}, ... ],   # up to 4
#    "totals":  ["TOTALS", "", "", "45,000.00", ...],           # optional, aligned to columns
#  }
# ══════════════════════════════════════════════════════════════════════════════
def generate_statement_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable, Image)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    DARK_GREEN  = colors.HexColor("#1A4731")
    MID_GREEN   = colors.HexColor("#2D6A4F")
    ACCENT      = colors.HexColor("#40916C")
    LIGHT_GREEN = colors.HexColor("#D8F3DC")
    LIGHT_GRAY  = colors.HexColor("#F2F2F2")
    WHITE       = colors.white
    BORDER_CLR  = colors.HexColor("#B0B0B0")

    def S(name, **kw): return ParagraphStyle(name, **kw)
    ST = {
        "company":    S("company",    fontName="Helvetica-Bold",    fontSize=18, textColor=WHITE,      alignment=TA_CENTER),
        "contact":    S("contact",    fontName="Helvetica-Oblique", fontSize=8,  textColor=WHITE,      alignment=TA_CENTER),
        "title":      S("title",      fontName="Helvetica-Bold",    fontSize=14, textColor=DARK_GREEN, alignment=TA_CENTER, spaceBefore=4, spaceAfter=4),
        "branch_tag": S("branch_tag", fontName="Helvetica-Bold",    fontSize=11, textColor=WHITE,      alignment=TA_CENTER),
        "meta_label": S("meta_label", fontName="Helvetica-Bold",    fontSize=8,  textColor=DARK_GREEN),
        "meta_value": S("meta_value", fontName="Helvetica",         fontSize=8,  textColor=colors.black),
        "sum_lbl":    S("sum_lbl",    fontName="Helvetica-Bold",    fontSize=7,  textColor=colors.HexColor("#555555"), alignment=TA_CENTER),
        "sum_val":    S("sum_val",    fontName="Helvetica-Bold",    fontSize=12, textColor=DARK_GREEN, alignment=TA_CENTER),
        "col_hdr":    S("col_hdr",    fontName="Helvetica-Bold",    fontSize=7.5,textColor=WHITE,      alignment=TA_CENTER, leading=9),
        "cell_l":     S("cell_l",     fontName="Helvetica",         fontSize=7.5,textColor=colors.black, alignment=TA_LEFT,   leading=9),
        "cell_c":     S("cell_c",     fontName="Helvetica",         fontSize=7.5,textColor=colors.black, alignment=TA_CENTER, leading=9),
        "cell_r":     S("cell_r",     fontName="Helvetica",         fontSize=7.5,textColor=colors.black, alignment=TA_RIGHT,  leading=9),
        "cell_l_red": S("cell_l_red", fontName="Helvetica-Bold",    fontSize=7.5,textColor=colors.HexColor("#C62828"), alignment=TA_LEFT,   leading=9),
        "cell_c_red": S("cell_c_red", fontName="Helvetica-Bold",    fontSize=7.5,textColor=colors.HexColor("#C62828"), alignment=TA_CENTER, leading=9),
        "cell_r_red": S("cell_r_red", fontName="Helvetica-Bold",    fontSize=7.5,textColor=colors.HexColor("#C62828"), alignment=TA_RIGHT,  leading=9),
        "total_lbl":  S("total_lbl",  fontName="Helvetica-Bold",    fontSize=8,  textColor=WHITE,      alignment=TA_CENTER),
        "total_val":  S("total_val",  fontName="Helvetica-Bold",    fontSize=8,  textColor=WHITE,      alignment=TA_RIGHT,  leading=9),
        "sig":        S("sig",        fontName="Helvetica-Oblique", fontSize=8,  textColor=colors.HexColor("#444444")),
        "footer":     S("footer",     fontName="Helvetica-Oblique", fontSize=7,  textColor=colors.HexColor("#888888"), alignment=TA_CENTER),
    }
    def P(text, style): return Paragraph(str(text), ST[style])

    title     = data.get("title", "STATEMENT")
    period    = data.get("period", "")
    branch    = data.get("branch", "All Branches")
    generated = data.get("generated", datetime.now().strftime("%d/%m/%Y, %H:%M:%S"))
    company   = data.get("company_name", "FAMOUSGATE HOTELS")
    address   = data.get("company_address", "Bomet, Kenya")
    email     = data.get("company_email", "famousgateshotelsbmt@gmail.com")
    phone     = data.get("company_phone", "0706 782 828")
    columns   = data.get("columns", [])
    rows      = data.get("rows", [])
    summary   = data.get("summary", [])
    totals    = data.get("totals")

    page_w, page_h = landscape(A4)
    margin = 12 * mm
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=margin)
    usable_w = page_w - 2 * margin
    story = []

    # ── HEADER BANNER with logo ───────────────────────────────────────────────
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        try:
            logo_cell = Image(LOGO_PATH, width=18*mm, height=18*mm)
        except Exception:
            logo_cell = P("FG", "company")
    logo_w = 22 * mm
    hdr = Table([[logo_cell, P(company, "company")]], colWidths=[logo_w, usable_w - logo_w])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), DARK_GREEN), ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",(0,0),(0,0), 4), ("TOPPADDING",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,-1),10),
    ]))
    contact = Table([[P(f"{address}   |   {email}   |   Tel: {phone}", "contact")]], colWidths=[usable_w])
    contact.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), MID_GREEN),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),5),
    ]))
    story += [hdr, contact, Spacer(1, 3*mm)]

    # ── TITLE + BRANCH TAG ────────────────────────────────────────────────────
    story.append(P(title, "title"))
    if branch and branch != "All Branches":
        bt = Table([[P(f"Branch: {branch}", "branch_tag")]], colWidths=[usable_w])
        bt.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), ACCENT),
            ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        story += [bt, Spacer(1, 2*mm)]
    story.append(HRFlowable(width=usable_w, thickness=2, color=DARK_GREEN, spaceAfter=3*mm))

    # ── META ──────────────────────────────────────────────────────────────────
    hw = usable_w / 2
    meta = Table([
        [P("Period:",    "meta_label"), P(period,           "meta_value"), P("Branch:",  "meta_label"), P(branch,          "meta_value")],
        [P("Generated:", "meta_label"), P(generated,        "meta_value"), P("Records:", "meta_label"), P(str(len(rows)),  "meta_value")],
    ], colWidths=[25*mm, hw-25*mm, 25*mm, hw-25*mm])
    meta.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,-1),LIGHT_GRAY), ("BACKGROUND",(2,0),(2,-1),LIGHT_GRAY),
        ("TOPPADDING",(0,0),(-1,-1),3), ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),4), ("GRID",(0,0),(-1,-1),0.4,BORDER_CLR),
    ]))
    story += [meta, Spacer(1, 3*mm)]

    # ── SUMMARY BOXES (up to 4) ───────────────────────────────────────────────
    if summary:
        sb = summary[:4]
        n = len(sb)
        qw = usable_w / n
        sum_tbl = Table([
            [P(b.get("label", ""), "sum_lbl") for b in sb],
            [P(b.get("value", ""), "sum_val") for b in sb],
        ], colWidths=[qw] * n)
        sum_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),LIGHT_GRAY), ("BACKGROUND",(0,1),(-1,1),WHITE),
            ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
            ("GRID",(0,0),(-1,-1),0.8,colors.HexColor("#555555")),
            ("LINEABOVE",(0,0),(-1,0),1.5,DARK_GREEN), ("LINEBELOW",(0,1),(-1,1),1.5,DARK_GREEN),
        ]))
        story += [sum_tbl, Spacer(1, 4*mm)]

    # ── MAIN TABLE ────────────────────────────────────────────────────────────
    aligns  = [c.get("align", "left") for c in columns]
    weights = [float(c.get("weight", 1) or 1) for c in columns]
    wsum    = sum(weights) or 1
    col_pts = [usable_w * (w / wsum) for w in weights]
    style_for = {"left": "cell_l", "center": "cell_c", "right": "cell_r"}

    table_data = [[P(c.get("header", ""), "col_hdr") for c in columns]]
    red_rows = []
    for r_idx, row in enumerate(rows):
        is_red = False
        if row:
            last_val = str(row[-1]).strip()
            if last_val in ('-', '0', '0.00', 'KES 0.00') or last_val.startswith('-'):
                is_red = True
            elif len(row) >= 4 and str(row[3]).strip() in ('-', '0', '0.00', 'Not set'):
                is_red = True
        if is_red:
            red_rows.append(r_idx + 1)

        cells = []
        for i, val in enumerate(row):
            al = aligns[i] if i < len(aligns) else "left"
            align_code = al[0] if al else "l"
            style_key = f"cell_{align_code}_red" if is_red else style_for.get(al, "cell_l")
            cells.append(P(val, style_key))
        table_data.append(cells)

    has_totals = bool(totals)
    if has_totals:
        table_data.append([P(v, "total_lbl" if i == 0 else "total_val")
                           for i, v in enumerate(totals)])

    n_rows = len(table_data)
    last = n_rows - 1
    data_last = last - 1 if has_totals else last
    tbl = Table(table_data, colWidths=col_pts, repeatRows=1)
    cmds = [
        ("BACKGROUND",(0,0),(-1,0),DARK_GREEN),
        ("GRID",(0,0),(-1,-1),0.4,BORDER_CLR),
        ("LINEBELOW",(0,0),(-1,0),1.0,WHITE),
        ("TOPPADDING",(0,0),(-1,-1),3), ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),3), ("RIGHTPADDING",(0,0),(-1,-1),3),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]
    if data_last >= 1:
        cmds.append(("ROWBACKGROUNDS",(0,1),(-1,data_last),[WHITE, LIGHT_GREEN]))
    for r_idx in red_rows:
        cmds.append(("BACKGROUND", (0, r_idx), (-1, r_idx), colors.HexColor("#FFEBEE")))
    if has_totals:
        cmds += [("BACKGROUND",(0,last),(-1,last),ACCENT),
                 ("LINEABOVE",(0,last),(-1,last),1.5,colors.HexColor("#555555"))]
    tbl.setStyle(TableStyle(cmds))
    story += [tbl, Spacer(1, 8*mm)]

    # ── SIGNATURES + FOOTER ───────────────────────────────────────────────────
    sig = Table([[P("Prepared by: "+"_"*38, "sig"), P("", "sig"), P("Verified by: "+"_"*38, "sig")]],
                colWidths=[usable_w*0.44, usable_w*0.12, usable_w*0.44])
    sig.setStyle(TableStyle([("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    story += [sig, Spacer(1, 4*mm),
              HRFlowable(width=usable_w, thickness=0.5, color=BORDER_CLR),
              Spacer(1, 2*mm),
              P(f"(c) 2026 {company} - Confidential Document   |   Generated: {generated}", "footer")]

    doc.build(story)
    buf.seek(0)
    return buf.read()
