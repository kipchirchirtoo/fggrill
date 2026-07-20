"""
Accounting Document Generator
Generates PDF and Excel exports for all accounting and bookkeeping documents
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfgen import canvas
from datetime import datetime
from decimal import Decimal
import io
import os
from typing import Dict, List, Any
from xml.sax.saxutils import escape
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class AccountingDocumentGenerator:
    """Generate professional accounting documents"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#333333'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))

    def _resolve_company_logo_path(self) -> str | None:
        """Resolve the shared FamousGate logo from the repo."""
        candidate = os.path.abspath(
            os.path.join(
                os.path.dirname(__file__),
                '..',
                '..',
                'frontend',
                'public',
                'fglogo.png',
            )
        )
        return candidate if os.path.exists(candidate) else None

    def _draw_company_header(self, canvas_obj, department_label: str):
        """Draw the standard company header with logo fallback."""
        logo_path = self._resolve_company_logo_path()
        text_x = inch

        if logo_path:
            try:
                canvas_obj.drawImage(
                    logo_path,
                    inch,
                    10.08 * inch,
                    width=0.72 * inch,
                    height=0.72 * inch,
                    preserveAspectRatio=True,
                    mask='auto',
                )
                text_x = inch + 0.9 * inch
            except Exception:
                text_x = inch

        canvas_obj.setFont('Helvetica-Bold', 16)
        canvas_obj.drawString(text_x, 10.52 * inch, "FamousGate Hotels")
        canvas_obj.setFont('Helvetica', 10)
        canvas_obj.drawString(text_x, 10.3 * inch, department_label)
        canvas_obj.line(inch, 10.18 * inch, 7.5 * inch, 10.18 * inch)
    
    def _add_header(self, canvas_obj, doc):
        """Add header to each page"""
        canvas_obj.saveState()
        self._draw_company_header(canvas_obj, "Accounting & Finance Department")
        canvas_obj.restoreState()
    
    def _add_footer(self, canvas_obj, doc):
        """Add footer to each page"""
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 8)
        canvas_obj.drawString(inch, 0.5*inch, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas_obj.drawRightString(7.5*inch, 0.5*inch, f"Page {doc.page}")
        canvas_obj.restoreState()
    
    def generate_trial_balance_pdf(self, data: Dict[str, Any]) -> bytes:
        """Generate Trial Balance PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5*inch, bottomMargin=inch)
        story = []
        
        # Title
        title = Paragraph("TRIAL BALANCE", self.styles['CustomTitle'])
        story.append(title)
        
        # Period info
        period_text = f"As of {data.get('as_of_date', datetime.now().strftime('%Y-%m-%d'))}"
        if data.get('branch_name'):
            period_text += f" - {data['branch_name']}"
        story.append(Paragraph(period_text, self.styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [['Account Code', 'Account Name', 'Debit (KES)', 'Credit (KES)']]
        
        total_debit = Decimal('0')
        total_credit = Decimal('0')
        
        for account in data.get('accounts', []):
            debit = Decimal(str(account.get('total_debit', 0)))
            credit = Decimal(str(account.get('total_credit', 0)))
            total_debit += debit
            total_credit += credit
            
            table_data.append([
                account.get('account_code', ''),
                account.get('account_name', ''),
                f"{debit:,.2f}" if debit > 0 else '-',
                f"{credit:,.2f}" if credit > 0 else '-'
            ])
        
        # Add totals
        table_data.append(['', 'TOTAL', f"{total_debit:,.2f}", f"{total_credit:,.2f}"])
        
        # Create table
        table = Table(table_data, colWidths=[1.2*inch, 3*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story, onFirstPage=self._add_header, onLaterPages=self._add_header)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_profit_loss_pdf(self, data: Dict[str, Any]) -> bytes:
        """Generate Profit & Loss Statement PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5*inch, bottomMargin=inch)
        story = []
        
        # Title
        title = Paragraph("PROFIT & LOSS STATEMENT", self.styles['CustomTitle'])
        story.append(title)
        
        # Period
        period = f"{data.get('start_date', '')} to {data.get('end_date', '')}"
        if data.get('branch_name'):
            period += f" - {data['branch_name']}"
        story.append(Paragraph(period, self.styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Revenue Section
        story.append(Paragraph("REVENUE", self.styles['CustomHeading']))
        revenue_data = [['Account', 'Amount (KES)']]
        total_revenue = Decimal('0')
        
        for item in data.get('revenue', []):
            amount = Decimal(str(item.get('amount', 0)))
            total_revenue += amount
            revenue_data.append([item.get('account_name', ''), f"{amount:,.2f}"])
        
        revenue_data.append(['Total Revenue', f"{total_revenue:,.2f}"])
        
        revenue_table = Table(revenue_data, colWidths=[4*inch, 2*inch])
        revenue_table.setStyle(self._get_financial_table_style())
        story.append(revenue_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Expenses Section
        story.append(Paragraph("EXPENSES", self.styles['CustomHeading']))
        expense_data = [['Account', 'Amount (KES)']]
        total_expenses = Decimal('0')
        
        for item in data.get('expenses', []):
            amount = Decimal(str(item.get('amount', 0)))
            total_expenses += amount
            expense_data.append([item.get('account_name', ''), f"{amount:,.2f}"])
        
        expense_data.append(['Total Expenses', f"{total_expenses:,.2f}"])
        
        expense_table = Table(expense_data, colWidths=[4*inch, 2*inch])
        expense_table.setStyle(self._get_financial_table_style())
        story.append(expense_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Net Profit/Loss
        net_profit = total_revenue - total_expenses
        net_color = colors.green if net_profit >= 0 else colors.red
        
        summary_data = [
            ['Total Revenue', f"{total_revenue:,.2f}"],
            ['Total Expenses', f"{total_expenses:,.2f}"],
            ['Net Profit/(Loss)', f"{net_profit:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('TEXTCOLOR', (1, -1), (1, -1), net_color),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
            ('TOPPADDING', (0, -1), (-1, -1), 10)
        ]))
        
        story.append(summary_table)
        
        doc.build(story, onFirstPage=self._add_header, onLaterPages=self._add_header)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_balance_sheet_pdf(self, data: Dict[str, Any]) -> bytes:
        """Generate Balance Sheet PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5*inch, bottomMargin=inch)
        story = []
        
        # Title
        title = Paragraph("BALANCE SHEET", self.styles['CustomTitle'])
        story.append(title)
        
        # Date
        as_of = f"As of {data.get('as_of_date', datetime.now().strftime('%Y-%m-%d'))}"
        if data.get('branch_name'):
            as_of += f" - {data['branch_name']}"
        story.append(Paragraph(as_of, self.styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Assets
        story.append(Paragraph("ASSETS", self.styles['CustomHeading']))
        assets_data = [['Account', 'Amount (KES)']]
        total_assets = Decimal('0')
        
        for item in data.get('assets', []):
            amount = Decimal(str(item.get('balance', 0)))
            total_assets += amount
            assets_data.append([item.get('account_name', ''), f"{amount:,.2f}"])
        
        assets_data.append(['Total Assets', f"{total_assets:,.2f}"])
        
        assets_table = Table(assets_data, colWidths=[4*inch, 2*inch])
        assets_table.setStyle(self._get_financial_table_style())
        story.append(assets_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Liabilities
        story.append(Paragraph("LIABILITIES", self.styles['CustomHeading']))
        liabilities_data = [['Account', 'Amount (KES)']]
        total_liabilities = Decimal('0')
        
        for item in data.get('liabilities', []):
            amount = Decimal(str(item.get('balance', 0)))
            total_liabilities += amount
            liabilities_data.append([item.get('account_name', ''), f"{amount:,.2f}"])
        
        liabilities_data.append(['Total Liabilities', f"{total_liabilities:,.2f}"])
        
        liabilities_table = Table(liabilities_data, colWidths=[4*inch, 2*inch])
        liabilities_table.setStyle(self._get_financial_table_style())
        story.append(liabilities_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Equity
        story.append(Paragraph("EQUITY", self.styles['CustomHeading']))
        equity_data = [['Account', 'Amount (KES)']]
        total_equity = Decimal('0')
        
        for item in data.get('equity', []):
            amount = Decimal(str(item.get('balance', 0)))
            total_equity += amount
            equity_data.append([item.get('account_name', ''), f"{amount:,.2f}"])
        
        equity_data.append(['Total Equity', f"{total_equity:,.2f}"])
        
        equity_table = Table(equity_data, colWidths=[4*inch, 2*inch])
        equity_table.setStyle(self._get_financial_table_style())
        story.append(equity_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Summary
        total_liabilities_equity = total_liabilities + total_equity
        summary_data = [
            ['Total Assets', f"{total_assets:,.2f}"],
            ['Total Liabilities + Equity', f"{total_liabilities_equity:,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('LINEABOVE', (0, 0), (-1, -1), 2, colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 10)
        ]))
        
        story.append(summary_table)
        
        doc.build(story, onFirstPage=self._add_header, onLaterPages=self._add_header)
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _get_financial_table_style(self):
        """Get standard financial table style"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')])
        ])
    
    def generate_journal_entries_excel(self, entries: List[Dict[str, Any]]) -> bytes:
        """Generate Journal Entries Excel export"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Journal Entries"
        
        # Headers
        headers = ['Date', 'Entry #', 'Description', 'Account', 'Debit', 'Credit', 'Reference', 'Created By']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for entry in entries:
            for item in entry.get('items', []):
                ws.append([
                    entry.get('entry_date', ''),
                    entry.get('entry_number', ''),
                    entry.get('description', ''),
                    item.get('account_name', ''),
                    item.get('debit', 0),
                    item.get('credit', 0),
                    entry.get('reference', ''),
                    entry.get('created_by_name', '')
                ])
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_supplier_statement_pdf(self, data: Dict[str, Any]) -> bytes:
        """Generate Supplier Statement PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5*inch, bottomMargin=inch)
        story = []
        
        # 1. Header Information
        supplier = data.get('supplier', {})
        
        # Statement Title
        story.append(Paragraph("SUPPLIER STATEMENT", self.styles['CustomTitle']))
        
        # Supplier Details & Period
        header_data = [
            [
                Paragraph(f"<b>To:</b><br/>{escape(str(supplier.get('name') or ''))}<br/>{escape(str(supplier.get('address') or ''))}<br/>{escape(str(supplier.get('email') or ''))}", self.styles['Normal']),
                Paragraph(f"<b>Statement Period:</b><br/>{escape(str(data.get('start_date', '')))} to {escape(str(data.get('end_date', '')))}<br/><br/><b>Currency:</b> KES", self.styles['Normal'])
            ]
        ]
        
        header_table = Table(header_data, colWidths=[4*inch, 3*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 2. Summary Section (Opening, Closing, Due)
        summary_data = [
            ['Opening Balance', 'Total Invoiced', 'Total Paid', 'Closing Balance'],
            [
                f"{(data.get('opening_balance') or 0):,.2f}",
                f"{(data.get('total_invoiced') or 0):,.2f}",
                f"{(data.get('total_paid') or 0):,.2f}",
                f"{(data.get('closing_balance') or 0):,.2f}"
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f8f9fa')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 3. Transactions Table
        trans_data = [['Date', 'Description', 'Ref', 'Debit', 'Credit', 'Balance']]
        
        # Add opening balance row
        trans_data.append([
            data.get('start_date', ''), 
            'Opening Balance', 
            '', 
            '', 
            '', 
            f"{data.get('opening_balance', 0):,.2f}"
        ])
        
        for trans in data.get('transactions', []):
            debit = Decimal(str(trans.get('debit_amount') or 0))
            credit = Decimal(str(trans.get('credit_amount') or 0))
            balance = Decimal(str(trans.get('running_balance') or 0))
            
            trans_data.append([
                trans.get('transaction_date', ''),
                Paragraph(escape(trans.get('description', '') or ''), self.styles['Normal']),
                trans.get('reference_number', ''),
                f"{debit:,.2f}" if debit > 0 else '-',
                f"{credit:,.2f}" if credit > 0 else '-',
                f"{balance:,.2f}"
            ])
            
        trans_table = Table(trans_data, colWidths=[1*inch, 2.5*inch, 1*inch, 0.9*inch, 0.9*inch, 1*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
        ]))
        story.append(trans_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 4. Aging Analysis
        aging = data.get('aging', {})
        story.append(Paragraph("Aging Analysis", self.styles['CustomHeading']))
        aging_data = [
            ['Current', '30 Days', '60 Days', '90+ Days', 'Total Due'],
            [
                f"{(aging.get('current_amount') or 0):,.2f}",
                f"{(aging.get('days_30_amount') or 0):,.2f}",
                f"{(aging.get('days_60_amount') or 0):,.2f}",
                f"{(aging.get('days_90_plus_amount') or 0):,.2f}",
                f"{(data.get('closing_balance') or 0):,.2f}"
            ]
        ]
        
        aging_table = Table(aging_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        aging_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, 1), (-1, 1), 1, colors.black),
            ('TEXTCOLOR', (4, 1), (4, 1), colors.red),
            ('FONTNAME', (4, 1), (4, 1), 'Helvetica-Bold'),
        ]))
        story.append(aging_table)
        
        # Build PDF
        doc.build(story, onFirstPage=self._add_header, onLaterPages=self._add_header)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_purchase_order_pdf(self, data: Dict[str, Any]) -> bytes:
        """Generate Purchase Order PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5*inch, bottomMargin=inch)
        story = []
        
        # 1. Title & key metadata
        title_table_data = [
            [
                Paragraph("PURCHASE ORDER", self.styles['CustomTitle']),
                Paragraph(f"<b>PO Number:</b> {escape(str(data.get('po_number', 'DRAFT')))}<br/><b>Date:</b> {escape(str(data.get('po_date', '')))}<br/><b>Status:</b> {escape(str(data.get('status', 'Draft').upper()))}", self.styles['Normal'])
            ]
        ]
        title_table = Table(title_table_data, colWidths=[4.5*inch, 2.5*inch])
        title_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        story.append(title_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 2. Supplier and Ship To
        supplier = data.get('supplier', {})
        org_details = [
            "<b>FamousGate Hotels</b>",
            "Bomet, Kenya",
            "Main Headquarters",
            "Tel: +254 706 782 828 | Email: info@famousgatehotels.com"
        ]
        
        address_data = [
            [
                Paragraph("<b>VENDOR:</b>", self.styles['Normal']),
                Paragraph("<b>SHIP TO:</b>", self.styles['Normal'])
            ],
            [
                Paragraph(f"{escape(str(supplier.get('name') or 'N/A'))}<br/>{escape(str(supplier.get('address') or 'N/A'))}<br/>Email: {escape(str(supplier.get('email') or 'N/A'))}<br/>PIN: {escape(str(supplier.get('supplier_pin') or 'N/A'))}", self.styles['Normal']),
                Paragraph("<b>Central Stores</b><br/>FamousGate Hotels<br/>Bomet, Kenya (HQ)<br/>Tel: +254 706 782 828", self.styles['Normal'])
            ]
        ]
        
        addr_table = Table(address_data, colWidths=[3.5*inch, 3.5*inch])
        addr_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LINEBELOW', (0, 0), (1, 0), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 10),
        ]))
        story.append(addr_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 3. Order Details (Terms)
        terms_data = [
            ['Payment Terms', 'Delivery Date', 'Delivery Method', 'Prepared By'],
            [
                data.get('payment_terms', '30 Days'),
                data.get('expected_delivery_date', 'Immediate'),
                data.get('delivery_terms', 'DDP'),
                data.get('created_by_name', 'System')
            ]
        ]
        terms_table = Table(terms_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        terms_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(terms_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 4. Items Table
        items_header = ['#', 'Item / Description', 'Unit', 'Qty', 'Unit Price', 'Total']
        items_data = [items_header]
        
        for idx, item in enumerate(data.get('items', []), 1):
            items_data.append([
                str(idx),
                Paragraph(escape(item.get('name', '') or ''), self.styles['Normal']),
                item.get('unit', 'Pcs'),
                str(item.get('quantity', 0)),
                f"{float(item.get('unit_price', 0)):,.2f}",
                f"{float(item.get('total_amount', 0) or (item.get('quantity', 0) * item.get('unit_price', 0))):,.2f}"
            ])
            
        # Fill empty rows if few items
        while len(items_data) < 8:
            items_data.append(['', '', '', '', '', ''])
            
        # Total Row
        total_amount = float(data.get('total_amount', 0))
        items_data.append(['', '', '', '', 'TOTAL (KES)', f"{total_amount:,.2f}"])
        
        items_table = Table(items_data, colWidths=[0.5*inch, 3*inch, 0.8*inch, 0.8*inch, 1*inch, 1.2*inch])
        
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('LINEBELOW', (0, -2), (-1, -2), 1, colors.black),
            ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (-2, -1), (-1, -1), colors.HexColor('#f8f9fa')),
        ]
        items_table.setStyle(TableStyle(table_style))
        story.append(items_table)
        story.append(Spacer(1, 0.3*inch))
        
        # 5. Authorization / Notes
        instructions = data.get('special_instructions', '')
        if instructions:
            story.append(Paragraph(f"<b>Special Instructions:</b><br/>{escape(instructions or '')}", self.styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
        story.append(Paragraph("This Purchase Order is an official document of FamousGate Hotels. Authorization is required for validity.", self.styles['Italic']))
        story.append(Spacer(1, 0.5*inch))
        
        # Signature block
        sig_data = [
            ['__________________________', '__________________________'],
            ['Prepared By', 'Authorized Signature']
        ]
        sig_table = Table(sig_data, colWidths=[3.5*inch, 3.5*inch])
        story.append(sig_table)

        doc.build(story, onFirstPage=self._add_header, onLaterPages=self._add_header)
        buffer.seek(0)
        return buffer.getvalue()

    # ──────────────────────────────────────────────────────────────────────────
    # EVENT ORDER
    # ──────────────────────────────────────────────────────────────────────────

    def _add_events_header(self, canvas_obj, doc):
        """Header / footer for event order pages."""
        canvas_obj.saveState()
        self._draw_company_header(canvas_obj, "Events & Catering Department")
        canvas_obj.setFont('Helvetica', 8)
        canvas_obj.drawString(inch, 0.5*inch, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        canvas_obj.drawRightString(7.5*inch, 0.5*inch, f"Page {doc.page}")
        canvas_obj.restoreState()

    def generate_event_order_pdf(self, data: Dict[str, Any]) -> bytes:
        """
        Generate a customer-facing Event Order PDF.

        Expected data keys:
          event_number, event_name, client_name, event_type
          event_date, pax, menu_package, charge_per_pax
          total_amount, amount_paid, payment_status, payment_method
          credit_due_date, notes, branch_name, created_by_name
          conference_halls  (dict with 'name', optional)
          items             (list of {description, unit, quantity, unit_price, total}, optional)
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            topMargin=1.5*inch, bottomMargin=inch,
            leftMargin=inch, rightMargin=inch
        )
        story = []

        # ── helpers ──────────────────────────────────────────────────────────
        EVENT_TYPE_LABELS = {
            'conference':      'CONFERENCE / MEETING',
            'buffet':          'BUFFET EVENT',
            'outside_catering':'OUTSIDE CATERING',
            'group_meal':      'GROUP MEAL',
        }
        PAYMENT_STATUS_LABELS = {
            'pending':      'Pending',
            'deposit_paid': 'Deposit Paid',
            'paid':         'Fully Paid',
            'credit':       'On Credit',
        }
        PAYMENT_METHOD_LABELS = {
            'cash':   'Cash',
            'mpesa':  'M-Pesa',
            'bank':   'Bank Transfer',
            'cheque': 'Cheque',
            'credit': 'Credit',
        }

        event_type_raw    = str(data.get('event_type', 'conference'))
        event_type_label  = EVENT_TYPE_LABELS.get(event_type_raw, event_type_raw.upper())
        payment_status    = PAYMENT_STATUS_LABELS.get(data.get('payment_status', 'pending'), 'Pending')
        payment_method    = PAYMENT_METHOD_LABELS.get(data.get('payment_method', 'cash'), 'Cash')

        issued_date = str(data.get('created_at', datetime.now().strftime('%Y-%m-%d')))[:10]

        # ── 1. Title / metadata row ───────────────────────────────────────────
        title_data = [[
            Paragraph("EVENT ORDER / FUNCTION CONFIRMATION", self.styles['CustomTitle']),
            Paragraph(
                f"<b>EO Number:</b> {escape(str(data.get('event_number', 'DRAFT')))}<br/>"
                f"<b>Date Issued:</b> {escape(issued_date)}<br/>"
                f"<b>Status:</b> {escape(payment_status)}",
                self.styles['Normal']
            )
        ]]
        title_table = Table(title_data, colWidths=[4.5*inch, 2.5*inch])
        title_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN',  (1, 0), (1,  0),  'RIGHT'),
        ]))
        story.append(title_table)

        # ── 2. Event-type banner ──────────────────────────────────────────────
        banner_data = [[Paragraph(f"<b>{event_type_label}</b>", self.styles['Normal'])]]
        banner_table = Table(banner_data, colWidths=[7*inch])
        banner_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR',     (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 11),
            ('TOPPADDING',    (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        story.append(Spacer(1, 0.15*inch))
        story.append(banner_table)
        story.append(Spacer(1, 0.25*inch))

        # ── 3. Client & Venue block ───────────────────────────────────────────
        hall_name = (data.get('conference_halls') or {}).get('name', '')
        hall_line = f"<br/><b>Hall:</b> {escape(hall_name)}" if hall_name else ''
        branch_name = escape(str(data.get('branch_name', 'FamousGate Hotels')))
        branch_address = escape(str(data.get('branch_address', 'FamousGate Hotels')))
        branch_phone = escape(str(data.get('branch_phone', '+254 706 782 828')))
        branch_email = escape(str(data.get('branch_email', 'info@famousgatehotels.com')))

        address_data = [
            [
                Paragraph("<b>CLIENT / BILL TO:</b>", self.styles['Normal']),
                Paragraph("<b>EVENT VENUE / FROM:</b>", self.styles['Normal'])
            ],
            [
                Paragraph(
                    f"<b>{escape(str(data.get('client_name', 'N/A')))}</b><br/>"
                    f"Event: {escape(str(data.get('event_name', 'N/A')))}",
                    self.styles['Normal']
                ),
                Paragraph(
                    f"<b>{branch_name}</b><br/>"
                    f"{branch_address}{hall_line}<br/>"
                    f"Tel: {branch_phone} | {branch_email}",
                    self.styles['Normal']
                )
            ]
        ]
        addr_table = Table(address_data, colWidths=[3.5*inch, 3.5*inch])
        addr_table.setStyle(TableStyle([
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('LINEBELOW',     (0, 0), (1,  0),  0.5, colors.grey),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ]))
        story.append(addr_table)
        story.append(Spacer(1, 0.3*inch))

        # ── 4. Event details strip ────────────────────────────────────────────
        pax = data.get('pax', 0)
        details_data = [
            ['Event Date', 'Pax Count', 'Package', 'Organized By'],
            [
                escape(str(data.get('event_date', 'TBD'))),
                str(pax),
                escape(str(data.get('menu_package', 'Standard'))),
                escape(str(data.get('created_by_name', data.get('created_by', 'Management')))),
            ]
        ]
        details_table = Table(details_data, colWidths=[1.75*inch, 1.25*inch, 2.5*inch, 1.5*inch])
        details_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#ecf0f1')),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE',      (0, 0), (-1, -1), 9),
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(details_table)
        story.append(Spacer(1, 0.3*inch))

        # ── 5. Services / menu items table ────────────────────────────────────
        items = list(data.get('items', []))

        # Auto-derive line items when the caller sends none
        if not items:
            charge_per_pax = float(data.get('charge_per_pax', 0))
            total_amount   = float(data.get('total_amount', 0))
            pkg_name       = data.get('menu_package', 'Event Package') or 'Event Package'
            pkg_total      = charge_per_pax * int(pax)
            items.append({
                'description': pkg_name,
                'unit':        'person',
                'quantity':    pax,
                'unit_price':  charge_per_pax,
                'total':       pkg_total,
            })
            remainder = total_amount - pkg_total
            if abs(remainder) > 0.01:
                items.append({
                    'description': 'Additional Charges / Miscellaneous',
                    'unit':        'lot',
                    'quantity':    1,
                    'unit_price':  remainder,
                    'total':       remainder,
                })

        col_header = ['#', 'Description / Service', 'Unit', 'Qty', 'Rate (KES)', 'Amount (KES)']
        rows = [col_header]

        for idx, item in enumerate(items, 1):
            rows.append([
                str(idx),
                Paragraph(
                    escape(str(item.get('description', item.get('name', '')))),
                    self.styles['Normal']
                ),
                escape(str(item.get('unit', 'person'))),
                str(item.get('quantity', 0)),
                f"{float(item.get('unit_price', item.get('rate', 0))):,.2f}",
                f"{float(item.get('total', item.get('total_amount', 0))):,.2f}",
            ])

        # Pad to at least 5 content rows so the table looks full
        while len(rows) < 6:
            rows.append(['', '', '', '', '', ''])

        total_amount = float(data.get('total_amount', 0))
        amount_paid  = float(data.get('amount_paid', 0))
        balance_due  = max(total_amount - amount_paid, 0)

        rows.append(['', '', '', '', 'SUBTOTAL (KES)',    f"{total_amount:,.2f}"])
        rows.append(['', '', '', '', 'AMOUNT PAID (KES)', f"{amount_paid:,.2f}"])
        rows.append(['', '', '', '', 'BALANCE DUE (KES)', f"{balance_due:,.2f}"])

        n = len(rows)
        col_widths = [0.4*inch, 2.85*inch, 0.7*inch, 0.6*inch, 1.2*inch, 1.25*inch]
        svc_table = Table(rows, colWidths=col_widths)

        svc_style = [
            # Header
            ('BACKGROUND',    (0, 0),    (-1, 0),    colors.HexColor('#2c3e50')),
            ('TEXTCOLOR',     (0, 0),    (-1, 0),    colors.whitesmoke),
            ('FONTNAME',      (0, 0),    (-1, 0),    'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0),    (-1, 0),    9),
            ('TOPPADDING',    (0, 0),    (-1, 0),    8),
            ('BOTTOMPADDING', (0, 0),    (-1, 0),    8),
            # Content rows
            ('FONTSIZE',      (0, 1),    (-1, n-4),  9),
            ('ALIGN',         (0, 0),    (-1, -1),   'LEFT'),
            ('ALIGN',         (2, 1),    (-1, n-4),  'RIGHT'),
            ('ALIGN',         (-2, n-3), (-1, -1),   'RIGHT'),
            # Grid on item rows only
            ('GRID',          (0, 0),    (-1, n-4),  0.5, colors.grey),
            # Divider above / below summary block
            ('LINEABOVE',     (0, n-3),  (-1, n-3),  1.5, colors.HexColor('#2c3e50')),
            ('LINEBELOW',     (0, n-1),  (-1, n-1),  1.5, colors.HexColor('#2c3e50')),
            # Subtotal row
            ('FONTNAME',      (-2, n-3), (-1, n-3),  'Helvetica-Bold'),
            ('BACKGROUND',    (-2, n-3), (-1, n-3),  colors.HexColor('#ecf0f1')),
            ('TOPPADDING',    (-2, n-3), (-1, n-3),  6),
            ('BOTTOMPADDING', (-2, n-3), (-1, n-3),  6),
            # Amount paid row
            ('BACKGROUND',    (-2, n-2), (-1, n-2),  colors.HexColor('#f8f9fa')),
            ('TOPPADDING',    (-2, n-2), (-1, n-2),  6),
            ('BOTTOMPADDING', (-2, n-2), (-1, n-2),  6),
            # Balance due row (dark highlight)
            ('FONTNAME',      (-2, n-1), (-1, n-1),  'Helvetica-Bold'),
            ('BACKGROUND',    (-2, n-1), (-1, n-1),  colors.HexColor('#2c3e50')),
            ('TEXTCOLOR',     (-2, n-1), (-1, n-1),  colors.whitesmoke),
            ('TOPPADDING',    (-2, n-1), (-1, n-1),  7),
            ('BOTTOMPADDING', (-2, n-1), (-1, n-1),  7),
        ]
        # Alternating row shading for item rows
        for i in range(2, n - 3, 2):
            svc_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa')))

        svc_table.setStyle(TableStyle(svc_style))
        story.append(svc_table)
        story.append(Spacer(1, 0.3*inch))

        # ── 6. Payment terms strip ────────────────────────────────────────────
        credit_due = str(data.get('credit_due_date', '') or '')
        pay_data = [
            ['Payment Method', 'Payment Status', 'Credit Due Date', ''],
            [
                payment_method,
                payment_status,
                escape(credit_due) if credit_due else '—',
                '',
            ]
        ]
        pay_table = Table(pay_data, colWidths=[1.75*inch, 1.75*inch, 1.75*inch, 1.75*inch])
        pay_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (2, 0), colors.HexColor('#ecf0f1')),
            ('FONTNAME',      (0, 0), (2, 0), 'Helvetica-Bold'),
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('GRID',          (0, 0), (2, -1), 0.5, colors.grey),
            ('FONTSIZE',      (0, 0), (-1, -1), 9),
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(pay_table)
        story.append(Spacer(1, 0.2*inch))

        # ── 7. Notes ─────────────────────────────────────────────────────────
        notes = str(data.get('notes', '') or '')
        if notes.strip():
            story.append(Paragraph(
                f"<b>Notes / Special Requirements:</b><br/>{escape(notes)}",
                self.styles['Normal']
            ))
            story.append(Spacer(1, 0.2*inch))

        # ── 8. Terms statement ────────────────────────────────────────────────
        story.append(Paragraph(
            "This Event Order is an official booking confirmation from FamousGate Hotels. "
            "Full payment is required before the event date unless credit terms have been "
            "agreed in writing. Cancellations must be communicated at least 72 hours in advance.",
            self.styles['Italic']
        ))
        story.append(Spacer(1, 0.5*inch))

        # ── 9. Signature block ────────────────────────────────────────────────
        sig_data = [
            ['__________________________', '__________________________'],
            ['Authorized (FamousGate Hotels)', 'Client Acceptance Signature'],
        ]
        sig_table = Table(sig_data, colWidths=[3.5*inch, 3.5*inch])
        sig_table.setStyle(TableStyle([
            ('ALIGN',    (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(sig_table)

        doc.build(story, onFirstPage=self._add_events_header, onLaterPages=self._add_events_header)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_event_order_pdf_v2(self, data: Dict[str, Any]) -> bytes:
        """Render the Event Order document with configured package/menu lines."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            topMargin=1.5 * inch,
            bottomMargin=inch,
            leftMargin=inch,
            rightMargin=inch,
        )
        story = []

        event_type_labels = {
            'conference': 'CONFERENCE / MEETING',
            'buffet': 'BUFFET EVENT',
            'outside_catering': 'OUTSIDE CATERING',
            'group_meal': 'GROUP MEAL',
        }
        payment_status_labels = {
            'pending': 'Pending',
            'deposit_paid': 'Deposit Paid',
            'paid': 'Fully Paid',
            'credit': 'On Credit',
        }
        payment_method_labels = {
            'cash': 'Cash',
            'mpesa': 'M-Pesa',
            'bank': 'Bank Transfer',
            'cheque': 'Cheque',
            'credit': 'Credit',
        }

        event_type_raw = str(data.get('event_type', 'conference'))
        event_type_label = event_type_labels.get(event_type_raw, event_type_raw.upper())
        payment_status = payment_status_labels.get(
            data.get('payment_status', 'pending'),
            'Pending',
        )
        payment_method = payment_method_labels.get(
            data.get('payment_method', 'cash'),
            'Cash',
        )
        issued_date = str(
            data.get('created_at', datetime.now().strftime('%Y-%m-%d'))
        )[:10]

        title_table = Table([[
            Paragraph(
                "EVENT ORDER / FUNCTION CONFIRMATION",
                self.styles['CustomTitle'],
            ),
            Paragraph(
                f"<b>EO Number:</b> {escape(str(data.get('event_number', 'DRAFT')))}<br/>"
                f"<b>Date Issued:</b> {escape(issued_date)}<br/>"
                f"<b>Status:</b> {escape(payment_status)}",
                self.styles['Normal'],
            ),
        ]], colWidths=[4.5 * inch, 2.5 * inch])
        title_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        story.append(title_table)

        banner_table = Table(
            [[Paragraph(f"<b>{event_type_label}</b>", self.styles['Normal'])]],
            colWidths=[7 * inch],
        )
        banner_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        story.append(Spacer(1, 0.15 * inch))
        story.append(banner_table)
        story.append(Spacer(1, 0.25 * inch))

        hall_name = (data.get('conference_halls') or {}).get('name', '')
        hall_line = f"<br/><b>Hall:</b> {escape(hall_name)}" if hall_name else ''
        branch_name = escape(str(data.get('branch_name', 'FamousGate Hotels')))
        branch_address = escape(str(data.get('branch_address', 'FamousGate Hotels')))
        branch_phone = escape(str(data.get('branch_phone', '+254 706 782 828')))
        branch_email = escape(
            str(data.get('branch_email', 'info@famousgatehotels.com'))
        )

        addr_table = Table([
            [
                Paragraph("<b>CLIENT / BILL TO:</b>", self.styles['Normal']),
                Paragraph("<b>EVENT VENUE / FROM:</b>", self.styles['Normal']),
            ],
            [
                Paragraph(
                    f"<b>{escape(str(data.get('client_name', 'N/A')))}</b><br/>"
                    f"Event: {escape(str(data.get('event_name', 'N/A')))}",
                    self.styles['Normal'],
                ),
                Paragraph(
                    f"<b>{branch_name}</b><br/>"
                    f"{branch_address}{hall_line}<br/>"
                    f"Tel: {branch_phone} | {branch_email}",
                    self.styles['Normal'],
                ),
            ],
        ], colWidths=[3.5 * inch, 3.5 * inch])
        addr_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LINEBELOW', (0, 0), (1, 0), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ]))
        story.append(addr_table)
        story.append(Spacer(1, 0.3 * inch))

        pax = int(float(data.get('pax', 0) or 0))
        details_table = Table([
            ['Event Date', 'Pax Count', 'Package', 'Organized By'],
            [
                escape(str(data.get('event_date', 'TBD'))),
                str(pax),
                escape(str(data.get('menu_package', 'Standard'))),
                escape(
                    str(data.get('created_by_name', data.get('created_by', 'Management')))
                ),
            ],
        ], colWidths=[1.75 * inch, 1.25 * inch, 2.5 * inch, 1.5 * inch])
        details_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(details_table)
        story.append(Spacer(1, 0.3 * inch))

        menu_lines = list(data.get('menu_lines', []))
        if menu_lines:
            story.append(
                Paragraph("<b>Configured Menu Package</b>", self.styles['Heading2'])
            )
            story.append(Spacer(1, 0.12 * inch))
            menu_rows = [['#', 'Menu Item', 'Qty / Pax', 'Planned Qty', 'Unit']]
            for idx, item in enumerate(menu_lines, 1):
                menu_rows.append([
                    str(idx),
                    Paragraph(
                        escape(
                            str(item.get('description', item.get('name', '')) or 'Menu item')
                        ),
                        self.styles['Normal'],
                    ),
                    f"{float(item.get('quantity_per_pax', 0) or 0):,.2f}",
                    f"{float(item.get('planned_total', 0) or 0):,.2f}",
                    escape(str(item.get('unit', 'pcs'))),
                ])

            menu_table = Table(
                menu_rows,
                colWidths=[0.4 * inch, 3.6 * inch, 1.0 * inch, 1.1 * inch, 0.9 * inch],
            )
            menu_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f1fb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f3b6d')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d3dce6')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]
            for i in range(2, len(menu_rows), 2):
                menu_style.append(
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fbff'))
                )
            menu_table.setStyle(TableStyle(menu_style))
            story.append(menu_table)
            story.append(Spacer(1, 0.25 * inch))

        items = list(data.get('items', []))
        if not items:
            charge_per_pax = float(data.get('charge_per_pax', 0) or 0)
            total_amount = float(data.get('total_amount', 0) or 0)
            pkg_name = data.get('menu_package', 'Event Package') or 'Event Package'
            pkg_total = charge_per_pax * pax
            items.append({
                'description': pkg_name,
                'unit': 'pax',
                'quantity': pax or 1,
                'unit_price': charge_per_pax or total_amount,
                'total': pkg_total or total_amount,
            })
            remainder = total_amount - (pkg_total or total_amount)
            if abs(remainder) > 0.01:
                items.append({
                    'description': 'Additional Charges / Adjustment',
                    'unit': 'lot',
                    'quantity': 1,
                    'unit_price': remainder,
                    'total': remainder,
                })

        rows = [['#', 'Description / Service', 'Unit', 'Qty', 'Rate (KES)', 'Amount (KES)']]
        for idx, item in enumerate(items, 1):
            rows.append([
                str(idx),
                Paragraph(
                    escape(str(item.get('description', item.get('name', '')))),
                    self.styles['Normal'],
                ),
                escape(str(item.get('unit', 'pax'))),
                str(item.get('quantity', 0)),
                f"{float(item.get('unit_price', item.get('rate', 0)) or 0):,.2f}",
                f"{float(item.get('total', item.get('total_amount', 0)) or 0):,.2f}",
            ])
        while len(rows) < 6:
            rows.append(['', '', '', '', '', ''])

        total_amount = float(data.get('total_amount', 0) or 0)
        amount_paid = float(data.get('amount_paid', 0) or 0)
        balance_due = max(total_amount - amount_paid, 0)
        rows.append(['', '', '', '', 'SUBTOTAL (KES)', f"{total_amount:,.2f}"])
        rows.append(['', '', '', '', 'AMOUNT PAID (KES)', f"{amount_paid:,.2f}"])
        rows.append(['', '', '', '', 'BALANCE DUE (KES)', f"{balance_due:,.2f}"])

        n = len(rows)
        svc_table = Table(
            rows,
            colWidths=[0.4 * inch, 2.85 * inch, 0.7 * inch, 0.6 * inch, 1.2 * inch, 1.25 * inch],
        )
        svc_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, n - 4), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, n - 4), 'RIGHT'),
            ('ALIGN', (-2, n - 3), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, n - 4), 0.5, colors.grey),
            ('LINEABOVE', (0, n - 3), (-1, n - 3), 1.5, colors.HexColor('#2c3e50')),
            ('LINEBELOW', (0, n - 1), (-1, n - 1), 1.5, colors.HexColor('#2c3e50')),
            ('FONTNAME', (-2, n - 3), (-1, n - 3), 'Helvetica-Bold'),
            ('BACKGROUND', (-2, n - 3), (-1, n - 3), colors.HexColor('#ecf0f1')),
            ('TOPPADDING', (-2, n - 3), (-1, n - 3), 6),
            ('BOTTOMPADDING', (-2, n - 3), (-1, n - 3), 6),
            ('BACKGROUND', (-2, n - 2), (-1, n - 2), colors.HexColor('#f8f9fa')),
            ('TOPPADDING', (-2, n - 2), (-1, n - 2), 6),
            ('BOTTOMPADDING', (-2, n - 2), (-1, n - 2), 6),
            ('FONTNAME', (-2, n - 1), (-1, n - 1), 'Helvetica-Bold'),
            ('BACKGROUND', (-2, n - 1), (-1, n - 1), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (-2, n - 1), (-1, n - 1), colors.whitesmoke),
            ('TOPPADDING', (-2, n - 1), (-1, n - 1), 7),
            ('BOTTOMPADDING', (-2, n - 1), (-1, n - 1), 7),
        ]
        for i in range(2, n - 3, 2):
            svc_style.append(
                ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa'))
            )
        svc_table.setStyle(TableStyle(svc_style))
        story.append(svc_table)
        story.append(Spacer(1, 0.3 * inch))

        credit_due = str(data.get('credit_due_date', '') or '')
        pay_table = Table([
            ['Payment Method', 'Payment Status', 'Credit Due Date', ''],
            [payment_method, payment_status, escape(credit_due) if credit_due else '—', ''],
        ], colWidths=[1.75 * inch, 1.75 * inch, 1.75 * inch, 1.75 * inch])
        pay_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (2, 0), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, 0), (2, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (2, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(pay_table)
        story.append(Spacer(1, 0.2 * inch))

        notes = str(data.get('notes', '') or '')
        if notes.strip():
            story.append(Paragraph(
                f"<b>Notes / Special Requirements:</b><br/>{escape(notes)}",
                self.styles['Normal'],
            ))
            story.append(Spacer(1, 0.2 * inch))

        story.append(Paragraph(
            "This Event Order confirms the package, pax, and charges agreed for the event. "
            "Kitchen production and stock issue will be fulfilled against the configured package items "
            "shown above. Full payment is required before the event date unless approved credit terms "
            "have been recorded in writing.",
            self.styles['Italic'],
        ))
        story.append(Spacer(1, 0.5 * inch))

        sig_table = Table([
            ['__________________________', '__________________________'],
            ['Authorized (FamousGate Hotels)', 'Client Acceptance Signature'],
        ], colWidths=[3.5 * inch, 3.5 * inch])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(sig_table)

        doc.build(story, onFirstPage=self._add_events_header, onLaterPages=self._add_events_header)
        buffer.seek(0)
        return buffer.getvalue()
