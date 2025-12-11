"""
FG Grill Report Generation Microservice
Generates PDF and Excel reports for branch oversight data
"""

import os
import io
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
CORS(app, origins=['http://localhost:3001', 'http://localhost:5000'])

# Configuration
API_BASE_URL = os.getenv('API_BASE_URL', 'http://localhost:5000')


def create_header_style():
    """Create custom header style for PDF reports"""
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a2e'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    return header_style


def create_subheader_style():
    """Create subheader style"""
    styles = getSampleStyleSheet()
    return ParagraphStyle(
        'SubHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#4a4a4a'),
        spaceBefore=15,
        spaceAfter=10
    )


def create_normal_style():
    """Create normal text style"""
    styles = getSampleStyleSheet()
    return ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333')
    )


def format_currency(amount):
    """Format amount as Kenyan Shillings"""
    return f"KES {amount:,.0f}"


def format_percentage(value):
    """Format value as percentage"""
    return f"{value:.1f}%"


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'report-generator',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/reports/performance', methods=['POST'])
def generate_performance_report():
    """Generate performance report in PDF or Excel format"""
    try:
        data = request.json
        format_type = request.args.get('format', 'pdf')
        period = request.args.get('period', 'week')
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if format_type == 'pdf':
            return generate_performance_pdf(data, period)
        elif format_type == 'excel':
            return generate_performance_excel(data, period)
        else:
            return jsonify({'error': 'Invalid format. Use pdf or excel'}), 400
            
    except Exception as e:
        app.logger.error(f"Error generating performance report: {str(e)}")
        return jsonify({'error': str(e)}), 500


def generate_performance_pdf(data, period):
    """Generate performance PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(
        f"Branch Performance Report - {period.capitalize()}",
        create_header_style()
    )
    elements.append(title)
    
    # Report info
    report_date = Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}",
        create_normal_style()
    )
    elements.append(report_date)
    elements.append(Spacer(1, 20))
    
    # Summary section
    summary_title = Paragraph("Executive Summary", create_subheader_style())
    elements.append(summary_title)
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Revenue', format_currency(data.get('total_revenue', 0))],
        ['Total Orders', f"{data.get('total_orders', 0):,}"],
        ['Average Satisfaction', f"{data.get('avg_satisfaction', 0):.1f}/5.0"],
        ['Best Performer', data.get('best_performer', 'N/A')],
        ['Worst Performer', data.get('worst_performer', 'N/A')]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Branch performance table
    branch_title = Paragraph("Branch Performance Details", create_subheader_style())
    elements.append(branch_title)
    
    branches = data.get('branches', [])
    if branches:
        branch_data = [
            ['Branch', 'Revenue', 'Change', 'Orders', 'Satisfaction', 'Target %']
        ]
        
        for branch in branches:
            revenue_change = float(branch.get('revenue_change', 0))
            change_str = f"+{revenue_change:.1f}%" if revenue_change > 0 else f"{revenue_change:.1f}%"
            
            branch_data.append([
                branch.get('branch_name', 'Unknown'),
                format_currency(branch.get('revenue', 0)),
                change_str,
                f"{branch.get('orders', 0):,}",
                f"{branch.get('customer_satisfaction', 0):.1f}/5.0",
                f"{branch.get('target_achievement', 0):.0f}%"
            ])
        
        branch_table = Table(branch_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch, 1.2*inch, 1*inch])
        branch_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(branch_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'performance-report-{period}-{datetime.now().strftime("%Y%m%d")}.pdf'
    )


def generate_performance_excel(data, period):
    """Generate performance Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"Branch Performance Report - {period.capitalize()}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    # Summary section
    ws['A4'] = "Executive Summary"
    ws['A4'].font = Font(bold=True, size=14)
    
    summary_data = [
        ('Total Revenue', format_currency(data.get('total_revenue', 0))),
        ('Total Orders', f"{data.get('total_orders', 0):,}"),
        ('Average Satisfaction', f"{data.get('avg_satisfaction', 0):.1f}/5.0"),
        ('Best Performer', data.get('best_performer', 'N/A')),
        ('Worst Performer', data.get('worst_performer', 'N/A'))
    ]
    
    for i, (label, value) in enumerate(summary_data, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Branch details
    start_row = 12
    ws[f'A{start_row}'] = "Branch Performance Details"
    ws[f'A{start_row}'].font = Font(bold=True, size=14)
    
    headers = ['Branch', 'Revenue', 'Change %', 'Orders', 'Satisfaction', 'Target %']
    header_row = start_row + 1
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    branches = data.get('branches', [])
    for row_idx, branch in enumerate(branches, start=header_row + 1):
        revenue_change = float(branch.get('revenue_change', 0))
        
        row_data = [
            branch.get('branch_name', 'Unknown'),
            format_currency(branch.get('revenue', 0)),
            f"{revenue_change:+.1f}%",
            branch.get('orders', 0),
            f"{branch.get('customer_satisfaction', 0):.1f}",
            f"{branch.get('target_achievement', 0):.0f}%"
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'performance-report-{period}-{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/api/reports/staff', methods=['POST'])
def generate_staff_report():
    """Generate staff report in PDF or Excel format"""
    try:
        data = request.json
        format_type = request.args.get('format', 'pdf')
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if format_type == 'pdf':
            return generate_staff_pdf(data)
        elif format_type == 'excel':
            return generate_staff_excel(data)
        else:
            return jsonify({'error': 'Invalid format'}), 400
            
    except Exception as e:
        app.logger.error(f"Error generating staff report: {str(e)}")
        return jsonify({'error': str(e)}), 500


def generate_staff_pdf(data):
    """Generate staff PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    
    # Title
    title = Paragraph("Staff Overview Report", create_header_style())
    elements.append(title)
    
    report_date = Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}",
        create_normal_style()
    )
    elements.append(report_date)
    elements.append(Spacer(1, 20))
    
    # Summary
    summary_title = Paragraph("Staff Summary", create_subheader_style())
    elements.append(summary_title)
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Staff', str(data.get('total_staff', 0))],
        ['Active Staff', str(data.get('active_staff', 0))],
        ['On Leave', str(data.get('on_leave', 0))],
        ['Avg Performance', f"{data.get('avg_performance', 0):.0f}%"],
        ['Avg Attendance', f"{data.get('avg_attendance', 0):.0f}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Branch summaries
    branch_title = Paragraph("Staff by Branch", create_subheader_style())
    elements.append(branch_title)
    
    branch_summaries = data.get('branch_summaries', [])
    if branch_summaries:
        branch_data = [['Branch', 'Total', 'Active', 'On Leave', 'Performance', 'Attendance']]
        
        for branch in branch_summaries:
            branch_data.append([
                branch.get('branch_name', 'Unknown'),
                str(branch.get('total_staff', 0)),
                str(branch.get('active', 0)),
                str(branch.get('on_leave', 0)),
                f"{branch.get('avg_performance', 0):.0f}%",
                f"{branch.get('avg_attendance', 0):.0f}%"
            ])
        
        branch_table = Table(branch_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1.2*inch, 1.2*inch])
        branch_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(branch_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'staff-report-{datetime.now().strftime("%Y%m%d")}.pdf'
    )


def generate_staff_excel(data):
    """Generate staff Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Overview"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Staff Overview Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
    
    # Summary
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=14)
    
    summary_items = [
        ('Total Staff', data.get('total_staff', 0)),
        ('Active Staff', data.get('active_staff', 0)),
        ('On Leave', data.get('on_leave', 0)),
        ('Avg Performance', f"{data.get('avg_performance', 0):.0f}%"),
        ('Avg Attendance', f"{data.get('avg_attendance', 0):.0f}%")
    ]
    
    for i, (label, value) in enumerate(summary_items, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Branch details
    start_row = 12
    ws[f'A{start_row}'] = "Staff by Branch"
    ws[f'A{start_row}'].font = Font(bold=True, size=14)
    
    headers = ['Branch', 'Total', 'Active', 'On Leave', 'Performance', 'Attendance']
    header_row = start_row + 1
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    branch_summaries = data.get('branch_summaries', [])
    for row_idx, branch in enumerate(branch_summaries, start=header_row + 1):
        row_data = [
            branch.get('branch_name', 'Unknown'),
            branch.get('total_staff', 0),
            branch.get('active', 0),
            branch.get('on_leave', 0),
            f"{branch.get('avg_performance', 0):.0f}%",
            f"{branch.get('avg_attendance', 0):.0f}%"
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'staff-report-{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/api/reports/compliance', methods=['POST'])
def generate_compliance_report():
    """Generate compliance report in PDF or Excel format"""
    try:
        data = request.json
        format_type = request.args.get('format', 'pdf')
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        if format_type == 'pdf':
            return generate_compliance_pdf(data)
        elif format_type == 'excel':
            return generate_compliance_excel(data)
        else:
            return jsonify({'error': 'Invalid format'}), 400
            
    except Exception as e:
        app.logger.error(f"Error generating compliance report: {str(e)}")
        return jsonify({'error': str(e)}), 500


def generate_compliance_pdf(data):
    """Generate compliance PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    
    # Title
    title = Paragraph("Compliance Overview Report", create_header_style())
    elements.append(title)
    
    report_date = Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}",
        create_normal_style()
    )
    elements.append(report_date)
    elements.append(Spacer(1, 20))
    
    # Summary
    summary_title = Paragraph("Compliance Summary", create_subheader_style())
    elements.append(summary_title)
    
    compliance_rate = data.get('overall_compliance_rate', 0)
    rate_color = colors.green if compliance_rate >= 90 else (colors.orange if compliance_rate >= 75 else colors.red)
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Requirements', str(data.get('total_requirements', 0))],
        ['Compliant', str(data.get('compliant', 0))],
        ['Non-Compliant', str(data.get('non_compliant', 0))],
        ['Pending', str(data.get('pending', 0))],
        ['Expired', str(data.get('expired', 0))],
        ['Overall Compliance Rate', f"{compliance_rate:.0f}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Branch compliance
    branch_title = Paragraph("Compliance by Branch", create_subheader_style())
    elements.append(branch_title)
    
    branch_summaries = data.get('branch_summaries', [])
    if branch_summaries:
        branch_data = [['Branch', 'Total', 'Compliant', 'Non-Compliant', 'Pending', 'Rate']]
        
        for branch in branch_summaries:
            branch_data.append([
                branch.get('branch_name', 'Unknown'),
                str(branch.get('total_requirements', 0)),
                str(branch.get('compliant', 0)),
                str(branch.get('non_compliant', 0)),
                str(branch.get('pending', 0)),
                f"{branch.get('compliance_rate', 0):.0f}%"
            ])
        
        branch_table = Table(branch_data, colWidths=[2*inch, 1*inch, 1.2*inch, 1.3*inch, 1*inch, 1*inch])
        branch_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(branch_table)
    
    # Critical issues
    critical_issues = data.get('critical_issues', [])
    if critical_issues:
        elements.append(Spacer(1, 20))
        critical_title = Paragraph("Critical Issues", create_subheader_style())
        elements.append(critical_title)
        
        critical_data = [['Branch', 'Category', 'Requirement', 'Status', 'Due Date']]
        for issue in critical_issues[:10]:  # Limit to 10
            critical_data.append([
                issue.get('branch_name', 'Unknown'),
                issue.get('category', 'N/A'),
                issue.get('requirement', 'N/A'),
                issue.get('status', 'N/A').replace('_', ' ').title(),
                issue.get('due_date', 'N/A')
            ])
        
        critical_table = Table(critical_data, colWidths=[1.5*inch, 1.5*inch, 2.5*inch, 1.2*inch, 1*inch])
        critical_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(critical_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'compliance-report-{datetime.now().strftime("%Y%m%d")}.pdf'
    )


def generate_compliance_excel(data):
    """Generate compliance Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compliance Summary"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    danger_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Compliance Overview Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
    
    # Summary
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=14)
    
    summary_items = [
        ('Total Requirements', data.get('total_requirements', 0)),
        ('Compliant', data.get('compliant', 0)),
        ('Non-Compliant', data.get('non_compliant', 0)),
        ('Pending', data.get('pending', 0)),
        ('Expired', data.get('expired', 0)),
        ('Overall Compliance Rate', f"{data.get('overall_compliance_rate', 0):.0f}%")
    ]
    
    for i, (label, value) in enumerate(summary_items, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Branch details
    start_row = 13
    ws[f'A{start_row}'] = "Compliance by Branch"
    ws[f'A{start_row}'].font = Font(bold=True, size=14)
    
    headers = ['Branch', 'Total', 'Compliant', 'Non-Compliant', 'Pending', 'Rate']
    header_row = start_row + 1
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    branch_summaries = data.get('branch_summaries', [])
    for row_idx, branch in enumerate(branch_summaries, start=header_row + 1):
        row_data = [
            branch.get('branch_name', 'Unknown'),
            branch.get('total_requirements', 0),
            branch.get('compliant', 0),
            branch.get('non_compliant', 0),
            branch.get('pending', 0),
            f"{branch.get('compliance_rate', 0):.0f}%"
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    # Critical issues sheet
    ws2 = wb.create_sheet("Critical Issues")
    ws2['A1'] = "Critical Issues"
    ws2['A1'].font = Font(bold=True, size=14)
    
    issue_headers = ['Branch', 'Category', 'Requirement', 'Status', 'Due Date', 'Notes']
    for col, header in enumerate(issue_headers, start=1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = danger_fill
        cell.border = border
    
    critical_issues = data.get('critical_issues', [])
    for row_idx, issue in enumerate(critical_issues, start=3):
        row_data = [
            issue.get('branch_name', 'Unknown'),
            issue.get('category', 'N/A'),
            issue.get('requirement', 'N/A'),
            issue.get('status', 'N/A').replace('_', ' ').title(),
            issue.get('due_date', 'N/A'),
            issue.get('notes', '')
        ]
        
        for col, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15
        ws2.column_dimensions[get_column_letter(col)].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'compliance-report-{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


if __name__ == '__main__':
    port = int(os.getenv('PORT', 5001))
    debug = os.getenv('FLASK_DEBUG', 'true').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
